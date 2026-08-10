#!/usr/bin/env python3
"""Genera SQL/CSV de códigos de barras desde el Excel homologado de tickets."""

from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(
    "/Users/ibarra/Library/CloudStorage/Dropbox/FarmaCapital/Tickets/"
    "FarmaCapital_inventario_homologado_completo.xlsx"
)
OUT_SQL = ROOT / "sql" / "actualizar_codigos_barras_tickets.sql"
OUT_PENDIENTES = ROOT / "sql" / "pendientes_codigo_barras.csv"
OUT_IMPORT_TEMPLATE = ROOT / "sql" / "captura_barcodes_manual.csv"


def sku_for_row(ticket: str, linea, nombre: str, barcode_raw) -> str:
    bc = re.sub(r"\D", "", str(barcode_raw or ""))
    if len(bc) >= 8:
        return f"FC-{bc[-8:]}"
    digest = hashlib.md5(f"{ticket}|{nombre}|{linea}".encode()).hexdigest()[:8].upper()
    return f"FC-{digest}"


def sql_quote(val: str) -> str:
    return "'" + val.replace("'", "''") + "'"


def main() -> None:
    xlsx = DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"No existe Excel: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Compras_maestro"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    updates: list[tuple[str, str, str, str]] = []
    pendientes: list[dict[str, str]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["N.º ticket / orden"]]:
            continue
        ticket = str(row[idx["N.º ticket / orden"]])
        linea = str(row[idx["Línea ticket"]])
        nombre = str(row[idx["Nombre / variante"]] or "").strip()
        lote = str(row[idx["Lote"]] or "").strip()
        bc_raw = row[idx["Código de barras"]]
        bc = re.sub(r"\D", "", str(bc_raw or ""))
        sku = sku_for_row(ticket, linea, nombre, bc_raw)

        if bc:
            updates.append((sku, bc, ticket, nombre))
        else:
            pendientes.append(
                {
                    "sku": sku,
                    "ticket": ticket,
                    "linea": linea,
                    "nombre": nombre,
                    "lote": lote,
                    "codigo_barras": "",
                }
            )

    # Unicidad barcode → conservar primera aparición
    seen_bc: set[str] = set()
    unique_updates: list[tuple[str, str, str, str]] = []
    dupes = 0
    for item in updates:
        sku, bc, ticket, nombre = item
        if bc in seen_bc:
            dupes += 1
            continue
        seen_bc.add(bc)
        unique_updates.append(item)

    lines = [
        "-- ============================================================",
        "-- FarmaCapital — Actualizar codigo_barras desde Excel homologado",
        "-- Ejecutar DESPUÉS de carga_inventario_tickets_EJECUTAR_1..4",
        "-- Solo filas con EAN en ticket (Bodega, Surtidor, FarmaLive).",
        f"-- Filas con barcode en Excel: {len(updates)} | UPDATE únicos: {len(unique_updates)}",
        "-- ============================================================",
        "",
        "begin;",
        "",
    ]

    for sku, bc, ticket, nombre in unique_updates:
        label = f"-- {ticket} {nombre[:60]}"
        lines.append(label)
        lines.append(
            f"update public.productos set codigo_barras = {sql_quote(bc)} "
            f"where sku = {sql_quote(sku)} "
            f"and (codigo_barras is null or btrim(codigo_barras) = '');"
        )
        lines.append("")

    lines.extend(
        [
            "commit;",
            "",
            "-- Verificación pistola POS (debe coincidir con filas con barcode en Excel)",
            "select count(*) as productos_con_barcode",
            "from public.productos",
            "where codigo_barras is not null and btrim(codigo_barras) <> '';",
            "",
        ]
    )

    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")

    with OUT_PENDIENTES.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["sku", "ticket", "linea", "nombre", "lote", "codigo_barras"],
        )
        w.writeheader()
        w.writerows(pendientes)

    with OUT_IMPORT_TEMPLATE.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["sku", "codigo_barras"])
        w.writeheader()
        for p in pendientes:
            w.writerow({"sku": p["sku"], "codigo_barras": ""})

    print(f"SQL updates: {OUT_SQL} ({len(unique_updates)} filas)")
    print(f"Pendientes sin EAN en ticket: {OUT_PENDIENTES} ({len(pendientes)} filas)")
    print(f"Plantilla captura manual: {OUT_IMPORT_TEMPLATE}")
    if dupes:
        print(f"(Omitidos {dupes} barcodes duplicados en Excel)")


if __name__ == "__main__":
    main()
