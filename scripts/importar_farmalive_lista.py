#!/usr/bin/env python3
"""Cruza la lista Farmalive (Excel) → producto_precios_referencia.

Solo matches por código de barras contra productos que ya existen.
Usa precio 2% (lista base). No usa especial ni día de descuento.

  python3 scripts/importar_farmalive_lista.py --dry-run
  python3 scripts/importar_farmalive_lista.py --apply
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "sql" / "pricing" / "generated"
DEFAULT_XLSX = Path(
    "/Users/ibarra/Desktop/LISTA DE PRECIOS  NORMAL 17 DE AGOSTO  2026.xlsx"
)
FECHA_LISTA = "2026-08-17"
FUENTE = "farmalive"
HOJA = "LISTA DE PRECIOS "


def cargar_env() -> dict[str, str]:
    out: dict[str, str] = {}
    env_path = ROOT / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            out[k.strip()] = v.strip().strip('"').strip("'")
    out.update({k: v for k, v in os.environ.items() if v})
    return out


def digits(val) -> str:
    return re.sub(r"\D", "", str(val or ""))


def barcode_keys(code: str) -> set[str]:
    d = digits(code)
    if not d:
        return set()
    keys = {d, d.lstrip("0") or "0"}
    if len(d) < 12:
        keys.add(d.zfill(12))
    if len(d) < 13:
        keys.add(d.zfill(13))
    if len(d) == 12:
        keys.add("0" + d)
    if len(d) == 13 and d.startswith("0"):
        keys.add(d[1:])
    return {k for k in keys if k}


def fetch_productos(url: str, key: str) -> list[dict]:
    headers = {"apikey": key, "Authorization": f"Bearer {key}"}
    rows: list[dict] = []
    offset = 0
    while True:
        h = {**headers, "Range": f"{offset}-{offset + 499}"}
        r = requests.get(
            f"{url}/rest/v1/productos",
            headers=h,
            params={
                "select": "id,sku,nombre,codigo_barras,activo,costo",
                "activo": "eq.true",
                "order": "id.asc",
            },
            timeout=60,
        )
        r.raise_for_status()
        batch = r.json()
        rows.extend(batch)
        if len(batch) < 500:
            break
        offset += 500
    return rows


def index_por_ean(productos: list[dict]) -> dict[str, list[dict]]:
    idx: dict[str, list[dict]] = {}
    for p in productos:
        for key in barcode_keys(p.get("codigo_barras") or ""):
            idx.setdefault(key, []).append(p)
    return idx


def leer_lista(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if HOJA not in wb.sheetnames:
        sys.exit(f"No está la hoja {HOJA!r}. Hojas: {wb.sheetnames}")
    ws = wb[HOJA]
    out: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 20 or not row:
            continue
        codigo, nombre = row[0], row[1]
        precio2 = row[5] if len(row) > 5 else None
        if codigo in (None, "") and not nombre:
            continue
        try:
            precio = float(precio2)
        except (TypeError, ValueError):
            continue
        if precio <= 0:
            continue
        nom = str(nombre or "").replace("_x000D_", " ").replace("\n", " ").strip()
        nom = re.sub(r"\s+", " ", nom)
        out.append({
            "ean": digits(codigo),
            "nombre": nom,
            "precio": round(precio, 2),
            "line": i,
        })
    wb.close()
    return out


def matchear(lista: list[dict], idx: dict[str, list[dict]]) -> tuple[list[dict], list[dict]]:
    matched: list[dict] = []
    unmatched: list[dict] = []
    seen_producto: set[int] = set()
    for row in lista:
        hits: list[dict] = []
        seen: set[int] = set()
        for key in barcode_keys(row["ean"]):
            for p in idx.get(key, []):
                if p["id"] not in seen:
                    seen.add(p["id"])
                    hits.append(p)
        if len(hits) != 1:
            unmatched.append(row)
            continue
        prod = hits[0]
        if prod["id"] in seen_producto:
            continue
        seen_producto.add(prod["id"])
        matched.append({
            "producto_id": prod["id"],
            "sku": prod.get("sku"),
            "nombre_catalogo": prod.get("nombre"),
            "costo": prod.get("costo"),
            "ean": row["ean"],
            "nombre_fuente": row["nombre"],
            "precio": row["precio"],
        })
    return matched, unmatched


def upsert_fuente(url: str, key: str) -> None:
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    r = requests.post(
        f"{url}/rest/v1/fuentes_precio",
        headers=headers,
        json=[{
            "id": FUENTE,
            "nombre": "Farmalive",
            "tipo": "compra",
            "metodo": "import_archivo",
            "notas": "Lista normal Club Iztapalapa. Precio base (2%), no campañas de día.",
        }],
        timeout=30,
    )
    if not r.ok:
        sys.exit(f"No se pudo registrar fuentes_precio: {r.status_code} {r.text[:240]}")


def apply_rest(url: str, key: str, matched: list[dict], archivo: str) -> None:
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    imp = requests.post(
        f"{url}/rest/v1/importaciones_referencia",
        headers=headers,
        json={
            "fuente": FUENTE,
            "tipo": "compra",
            "fecha_lista": FECHA_LISTA,
            "archivo": archivo,
            "filas_ok": len(matched),
            "filas_error": 0,
            "notas": "importar_farmalive_lista.py · precio 2% · match EAN",
        },
        timeout=60,
    )
    imp.raise_for_status()
    import_id = imp.json()[0]["id"]

    for i in range(0, len(matched), 80):
        chunk = matched[i : i + 80]
        payload = [
            {
                "producto_id": m["producto_id"],
                "fuente": FUENTE,
                "tipo": "compra",
                "precio": m["precio"],
                "fecha": FECHA_LISTA,
                "nombre_fuente": m["nombre_fuente"],
                "sku_externo": m["ean"],
                "confianza": 100,
                "origen": "import_csv",
                "import_id": import_id,
                "notas": "lista 17-ago-2026 · precio 2%",
            }
            for m in chunk
        ]
        r = requests.post(
            f"{url}/rest/v1/producto_precios_referencia",
            headers=headers,
            json=payload,
            timeout=120,
        )
        r.raise_for_status()
        print(f"  Insertadas {min(i + 80, len(matched))}/{len(matched)}")
        time.sleep(0.15)


def sql_quote(s: str) -> str:
    return "'" + str(s).replace("'", "''") + "'"


def generate_sql(matched: list[dict], archivo: str) -> str:
    lines = [
        f"-- Farmalive lista {FECHA_LISTA} · {len(matched)} matches EAN · precio 2%",
        f"-- Archivo: {archivo}",
        "",
        "begin;",
        "",
        "insert into public.fuentes_precio (id, nombre, tipo, metodo, notas) values",
        "  ('farmalive', 'Farmalive', 'compra', 'import_archivo',",
        "   'Lista normal Club Iztapalapa. Precio base (2%), no campañas de día.')",
        "on conflict (id) do update set",
        "  nombre = excluded.nombre,",
        "  tipo = excluded.tipo,",
        "  metodo = excluded.metodo,",
        "  notas = excluded.notas;",
        "",
        "with imp as (",
        "  insert into public.importaciones_referencia (fuente, tipo, fecha_lista, archivo, filas_ok, notas)",
        f"  values ('farmalive', 'compra', '{FECHA_LISTA}', {sql_quote(archivo)}, {len(matched)},",
        "          'importar_farmalive_lista.py · precio 2% · match EAN')",
        "  returning id",
        ")",
        "insert into public.producto_precios_referencia (",
        "  producto_id, fuente, tipo, precio, fecha, nombre_fuente, sku_externo,",
        "  confianza, origen, import_id, notas",
        ")",
        "select",
        f"  v.producto_id, 'farmalive', 'compra', v.precio, '{FECHA_LISTA}'::date,",
        "  v.nombre_fuente, v.sku_externo, 100, 'import_csv', imp.id,",
        "  'lista 17-ago-2026 · precio 2%'",
        "from imp, (values",
    ]
    value_rows = []
    for m in matched:
        value_rows.append(
            f"  ({m['producto_id']}::bigint, {m['precio']}::numeric, "
            f"{sql_quote(m['nombre_fuente'])}, {sql_quote(m['ean'])})"
        )
    lines.append(",\n".join(value_rows))
    lines.extend([
        ") as v(producto_id, precio, nombre_fuente, sku_externo);",
        "",
        "commit;",
        "",
    ])
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Importar lista Farmalive a referencias de compra")
    parser.add_argument("--archivo", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sql-only", action="store_true")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    if not args.archivo.exists():
        sys.exit(f"No existe: {args.archivo}")

    env = cargar_env()
    url = env.get("REACT_APP_SUPABASE_URL") or env.get("SUPABASE_URL") or ""
    key = (
        env.get("SUPABASE_SERVICE_ROLE_KEY")
        or env.get("REACT_APP_SUPABASE_ANON_KEY")
        or ""
    )
    if not url or not key:
        sys.exit("Falta URL/key de Supabase en .env")

    print("Catálogo Supabase…")
    productos = fetch_productos(url, key)
    print(f"  {len(productos)} activos")
    idx = index_por_ean(productos)

    print(f"Leyendo {args.archivo.name}…")
    lista = leer_lista(args.archivo)
    print(f"  {len(lista)} renglones con precio 2%")

    matched, unmatched = matchear(lista, idx)
    mas_barato = 0
    for m in matched:
        try:
            costo = float(m["costo"]) if m.get("costo") is not None else None
        except (TypeError, ValueError):
            costo = None
        if costo and m["precio"] < costo - 0.05:
            mas_barato += 1

    print(f"Matches EAN: {len(matched)}")
    print(f"Sin match (no están en catálogo): {len(unmatched)}")
    print(f"Más baratos que tu costo: {mas_barato}")
    print("Ejemplos:")
    for m in matched[:10]:
        print(f"  {m['sku']}  ${m['precio']:.2f}  {m['nombre_catalogo'][:42]}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_sql = OUT_DIR / f"import_referencias_farmalive_{FECHA_LISTA.replace('-', '')}.sql"
    out_sql.write_text(generate_sql(matched, args.archivo.name), encoding="utf-8")
    print(f"SQL: {out_sql}")

    if args.dry_run or (not args.apply and not args.sql_only):
        if not args.apply:
            print("Para cargar: python3 scripts/importar_farmalive_lista.py --apply")
        return 0

    if args.sql_only:
        return 0

    print("Aplicando…")
    upsert_fuente(url, key)
    apply_rest(url, key, matched, args.archivo.name)
    print("Listo.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
