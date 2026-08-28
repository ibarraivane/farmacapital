#!/usr/bin/env python3
"""Cruza el Excel de Similares (articulos_farmacias.xlsx) contra el inventario vivo.

Produce qué genéricos de Similares ya tienes, cuáles faltan, y cuánto pedir
para un stock tipo sucursal pequeña (no el almacén completo de la cadena).
"""
from __future__ import annotations

import importlib.util
import json
import math
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "pricing" / "fuentes" / "articulos_farmacias.xlsx"
OUT_XLSX = ROOT / "pricing" / "reportes" / "cruce_similares_pedido_20260820.xlsx"
OUT_JSON = Path("/tmp/fc_cruce_similares.json")
LEVIC = Path("/Users/ibarra/Downloads/Pedidos_FarmaCapital_20260819 (1).xlsx")

spec = importlib.util.spec_from_file_location(
    "sync_excel", ROOT / "scripts" / "sync_precios_excel_farmacias.py"
)
ex = importlib.util.module_from_spec(spec)
spec.loader.exec_module(ex)
sync = ex.sync
sync.SINONIMOS.setdefault("cla", "clavulanico")
sync.SINONIMOS.setdefault("clavulanico", "clavulanico")
sync.SINONIMOS.setdefault("clavulanico", "clavulanico")
sync.SINONIMOS.setdefault("tioctico", "tioctico")

UMBRAL_CUBIERTO = 80
UMBRAL_EQUIV = 70

CLASE_A = {
    "paracetamol", "ibuprofeno", "naproxeno", "diclofenaco", "ketorolaco", "metamizol",
    "omeprazol", "pantoprazol", "ranitidina",
    "metformina", "glibenclamida",
    "losartan", "amlodipino", "enalapril", "captopril",
    "amoxicilina", "ampicilina", "ciprofloxacino", "metronidazol", "trimetoprima",
    "azitromicina",
    "loratadina", "cetirizina", "clorfenamina",
    "atorvastatina",
    "acetilsalicilico",
    "salbutamol", "ambroxol",
    "clotrimazol",
    "furosemida",
    "sildenafil",
}

MARCA_PROPIA_SIMI = re.compile(
    r"\b(simi\s?flex|simi\s?fx|simiflex|simibacil|simipro|simifibra|simifila|"
    r"simialoe|dr simi|multigomi|gomitas|nopal/chia)\b",
    re.I,
)

SKIP_JER = {"SOUVENIRS", "PERFUMERIA", "ALIMENTOS", "MATERIALES DIVERSOS", "PROMOCIONALES"}
SKIP_LINEA = {"SOUVENIRS"}


FORMA_FAMILIA = {
    "tableta": "oral_solido",
    "capsula": "oral_solido",
    "comprimido": "oral_solido",
    "gragea": "oral_solido",
    "suspension": "oral_liq",
    "jarabe": "oral_liq",
    "solucion": "oral_liq",
    "polvo": "oral_liq",
    "crema": "topico",
    "gel": "topico",
    "unguento": "topico",
    "pomada": "topico",
}


def formas_compatibles(a: str | None, b: str | None) -> bool:
    if not a or not b:
        return True
    if a == b:
        return True
    return FORMA_FAMILIA.get(a) == FORMA_FAMILIA.get(b) and FORMA_FAMILIA.get(a) is not None


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.lower()).strip()
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.lower()).strip()


def activos_sim(desc: str, concentracion: str) -> list[str]:
    """Moléculas del genérico Similares: cabeza de cada segmento antes de la dosis."""
    cabeza = re.split(r"\d", desc or "", maxsplit=1)[0]
    return sync.ingredientes(cabeza.replace("/", " + "))


def clase_y_stock(sim: dict) -> tuple[str, int, int]:
    """(clase, minimo sucursal, objetivo sucursal). Pedido inicial = minimo."""
    forma = (sim.get("forma") or "")
    jer = sim.get("jerarquia") or ""
    precio = float(sim.get("precio") or 0)
    desc = " ".join([sim.get("descripcion") or "", sim.get("marca") or ""])
    nd = norm(desc)

    if MARCA_PROPIA_SIMI.search(desc) or jer == "SUPLEMENTOS":
        return "C", 2, 3
    if precio >= 250:
        return "C", 1, 2
    if precio >= 120:
        return "B", 2, 4
    if "inyect" in (forma or "") or "ampolleta" in nd or "ampolla" in nd:
        return "INY", 2, 3
    if jer == "MATERIAL DE CURACION":
        return "CUR", 6, 10
    if any(tok in nd for tok in CLASE_A):
        return "A", 6, 10
    linea = sim.get("linea") or ""
    if linea in ("ANALGESICOS", "DIABETES", "CARDIOVASCULARES", "ESTOMACALES (GASTRO)", "ANTIHISTAMINICOS"):
        return "A", 5, 8
    if linea in ("ANTIBIOTICOS", "RESPIRATORIOS", "ANTIMICOTICOS"):
        return "B", 3, 6
    return "C", 2, 4


def cargar_similares_unicos(ruta: Path) -> list[dict]:
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_sku: dict[str, dict] = {}
    for fila in ws.iter_rows(min_row=3, values_only=True):
        sku = str(fila[0] or "").strip()
        if not sku or sku == "None":
            continue
        estatus = str(fila[13] or "").strip().upper() if len(fila) > 13 else "ACTIVO"
        if estatus and estatus not in ("ACTIVO", ""):
            continue
        marca = (fila[1] or "").strip() if isinstance(fila[1], str) else ""
        desc = (fila[2] or "").strip() if isinstance(fila[2], str) else ""
        if not desc:
            continue
        try:
            precio = float(str(fila[3]).replace("$", "").replace(",", ""))
        except (TypeError, ValueError):
            precio = 0.0
        concentracion = str(fila[4] or "").strip()
        contenido = str(fila[5] or "").strip()
        presentacion = str(fila[7] or "").strip() if len(fila) > 7 else ""
        linea = str(fila[9] or "").strip() if len(fila) > 9 else ""
        jerarquia = str(fila[10] or "").strip() if len(fila) > 10 else ""
        grupo = str(fila[11] or "").strip() if len(fila) > 11 else ""
        if jerarquia in SKIP_JER or linea in SKIP_LINEA:
            continue
        if jerarquia not in ("MEDICAMENTOS", "MATERIAL DE CURACION", "SUPLEMENTOS", "SALUD SEXUAL", "MATERNIDAD", "REMEDIOS HERBOLARIOS"):
            continue
        cand = {
            "sim_sku": sku,
            "marca": marca,
            "marca_norm": sync.normalizar(marca),
            "descripcion": desc,
            "desc_norm": sync.normalizar(desc),
            "precio": precio,
            "concentracion": concentracion,
            "contenido": contenido,
            "presentacion": presentacion,
            "linea": linea,
            "jerarquia": jerarquia,
            "grupo": grupo,
            "concentraciones": sync.extraer_concentraciones(f"{concentracion} {desc}"),
            "cantidad": sync.extraer_cantidad(contenido) or sync.extraer_cantidad(desc),
            "forma": sync.extraer_forma(presentacion) or sync.extraer_forma(desc),
            "ingredientes": activos_sim(desc, concentracion),
            "n_marcas": 1,
        }
        prev = by_sku.get(sku)
        if prev is None:
            by_sku[sku] = cand
        else:
            prev["n_marcas"] += 1
            if len(desc) > len(prev["descripcion"]):
                cand["n_marcas"] = prev["n_marcas"]
                by_sku[sku] = cand
    wb.close()
    return list(by_sku.values())


def fetch_productos_stock(url: str, key: str) -> list[dict]:
    import urllib.request

    H = {"apikey": key, "Authorization": f"Bearer {key}"}
    out = []
    start = 0
    while True:
        req = urllib.request.Request(
            url + "/rest/v1/productos?select=id,sku,nombre,principio_activo,marca,presentacion,forma_farmaceutica,concentracion,costo,precio,categoria,tipo,stock,stock_minimo,activo&activo=eq.true&order=nombre",
            headers={**H, "Range": f"{start}-{start + 999}"},
        )
        with urllib.request.urlopen(req, timeout=90) as r:
            chunk = json.loads(r.read())
        out.extend(chunk)
        if len(chunk) < 1000:
            break
        start += 1000
    return out


def cargar_pedido_levic(path: Path) -> dict[str, int]:
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Levic"]
    incoming: dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 6:
            continue
        sku = str(row[1] or "").strip()
        try:
            qty = int(row[3] or 0)
        except (TypeError, ValueError):
            continue
        if sku:
            incoming[sku] = incoming.get(sku, 0) + qty
    wb.close()
    return incoming


def token_index(productos: list[dict], atts: list[dict]) -> dict[str, list[int]]:
    idx: dict[str, list[int]] = defaultdict(list)
    for i, (p, att) in enumerate(zip(productos, atts)):
        toks = set(sync.tokens_utiles(att["texto"]))
        toks.update(att["ingredientes"])
        for t in toks:
            if len(t) >= 4:
                idx[t].append(i)
    return idx


def main() -> None:
    env = sync.cargar_env()
    url = env.get("REACT_APP_SUPABASE_URL", "").rstrip("/")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY") or env.get("REACT_APP_SUPABASE_ANON_KEY", "")
    if not url or not key:
        sys.exit("Faltan credenciales")

    sim = cargar_similares_unicos(EXCEL)
    print(f"Similares únicos (medicamentos/curación/suplementos): {len(sim)}")
    productos = fetch_productos_stock(url, key)
    print(f"Inventario activo: {len(productos)}")
    incoming = cargar_pedido_levic(LEVIC)
    print(f"Pedido Levic 19-ago: {len(incoming)} SKUs, {sum(incoming.values())} pzas")

    # fake catalog list for vagos — use all similar rows' descriptions
    vagos = set()
    from collections import Counter
    freq: Counter[str] = Counter()
    for c in sim:
        freq.update(set(sync.tokens_utiles(c["descripcion"])))
    minimo = max(25, int(len(sim) * 0.02))
    vagos = {t for t, n in freq.items() if n >= minimo}

    atts = []
    for p in productos:
        att = ex.nuestros_atributos(p)
        extra = " ".join(filter(None, [str(p.get("concentracion") or ""), str(p.get("forma_farmaceutica") or "")]))
        if extra:
            att["texto"] = (att["texto"] + " " + extra).strip()
            att["concentraciones"] |= sync.extraer_concentraciones(extra)
            att["cantidad"] = att["cantidad"] or sync.extraer_cantidad(att["texto"])
            att["forma"] = att["forma"] or sync.extraer_forma(extra)
        atts.append(att)
    idx = token_index(productos, atts)

    covered: dict[str, dict] = {}
    for s in sim:
        toks = set(sync.tokens_utiles(s["descripcion"]))
        toks.update(s.get("ingredientes") or [])
        cand_i = set()
        for t in toks:
            if len(t) >= 4:
                cand_i.update(idx.get(t, []))
        best = None
        best_score = 0
        best_raz = []
        best_equiv = False
        for i in cand_i:
            score, razones, _ = ex.evaluar(atts[i], s, vagos)
            sim_ing = [sync.canonico(x) for x in (s.get("ingredientes") or []) if x not in vagos and len(x) >= 4]
            fc_ing = [sync.canonico(x) for x in atts[i]["ingredientes"]]
            fc_txt = " ".join(atts[i]["ingredientes"] + sync.tokens_utiles(atts[i]["texto"]))
            faltan = []
            if sim_ing:
                faltan = [x for x in sim_ing if x not in fc_ing and x not in fc_txt]
                if faltan:
                    score = min(score, 55)
                    razones = razones + [f"faltan en tu ficha: {', '.join(faltan)}"]
                if len(sim_ing) >= 2 and len(fc_ing) == 1:
                    score = min(score, 58)
                    razones = razones + ["combinado Similares vs monofármaco tuyo"]
            pa_ok = bool(sim_ing) and not faltan
            conc_ok = bool(atts[i]["concentraciones"] & s["concentraciones"]) if (atts[i]["concentraciones"] and s["concentraciones"]) else False
            forma_ok = formas_compatibles(atts[i]["forma"], s["forma"])
            equiv = pa_ok and conc_ok and forma_ok
            if equiv and score < UMBRAL_EQUIV:
                score = max(score, UMBRAL_EQUIV)
            if score > best_score:
                best_score = score
                best = productos[i]
                best_raz = razones
                best_equiv = equiv
        if best is not None and (best_score >= UMBRAL_CUBIERTO or best_equiv):
            covered[s["sim_sku"]] = {
                "score": best_score,
                "razones": "; ".join(best_raz),
                "prod": best,
                "equiv": best_equiv,
            }

    huecos = []
    rellenar = []
    ok = []
    for s in sim:
        clase, minimo_s, objetivo = clase_y_stock(s)
        hit = covered.get(s["sim_sku"])
        rec = {
            **s,
            "clase": clase,
            "minimo": minimo_s,
            "objetivo": objetivo,
        }
        if not hit:
            rec.update({
                "estado": "HUECO",
                "fc_sku": "",
                "fc_nombre": "",
                "stock": 0,
                "incoming": 0,
                "pedir": minimo_s,
                "match": 0,
                "razones": "",
            })
            huecos.append(rec)
            continue
        p = hit["prod"]
        stock = int(p.get("stock") or 0)
        inc = int(incoming.get(p.get("sku") or "", 0))
        disponible = stock + inc
        rec.update({
            "fc_sku": p.get("sku") or "",
            "fc_nombre": p.get("nombre") or "",
            "stock": stock,
            "incoming": inc,
            "match": hit["score"],
            "razones": hit["razones"],
        })
        if disponible < minimo_s:
            rec["estado"] = "RELLENAR"
            rec["pedir"] = minimo_s - disponible
            rellenar.append(rec)
        else:
            rec["estado"] = "OK"
            rec["pedir"] = 0
            ok.append(rec)

    def prio_key(r):
        order = {"A": 0, "INY": 1, "CUR": 2, "B": 3, "C": 4}
        return (order.get(r["clase"], 9), -(r.get("precio") or 0), r["descripcion"])

    huecos.sort(key=prio_key)
    rellenar.sort(key=prio_key)

    resumen = {
        "similares_unicos": len(sim),
        "inventario_activo": len(productos),
        "cubiertos": len(ok) + len(rellenar),
        "ok": len(ok),
        "rellenar": len(rellenar),
        "huecos": len(huecos),
        "huecos_a": sum(1 for r in huecos if r["clase"] == "A"),
        "huecos_b": sum(1 for r in huecos if r["clase"] == "B"),
        "huecos_c": sum(1 for r in huecos if r["clase"] == "C"),
        "huecos_iny": sum(1 for r in huecos if r["clase"] == "INY"),
        "pzas_huecos": sum(r["pedir"] for r in huecos),
        "pzas_rellenar": sum(r["pedir"] for r in rellenar),
        "pzas_huecos_ab": sum(r["pedir"] for r in huecos if r["clase"] in ("A", "B")),
        "cobertura_pct": round(100 * (len(ok) + len(rellenar)) / len(sim), 1) if sim else 0,
    }
    print(json.dumps(resumen, indent=2))
    print("huecos A sample:")
    for r in huecos:
        if r["clase"] == "A":
            print(f"  {r['sim_sku']:>6} {r['pedir']}u ${r['precio']:>7.0f}  {r['descripcion'][:55]} | {r['contenido']}")

    OUT_JSON.write_text(json.dumps({
        "resumen": resumen,
        "huecos": huecos,
        "rellenar": rellenar,
        "ok_n": len(ok),
    }, ensure_ascii=False, default=str))
    print("json", OUT_JSON)

    escribir_xlsx(resumen, huecos, rellenar, ok)
    print("xlsx", OUT_XLSX)


def escribir_xlsx(resumen, huecos, rellenar, ok):
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    fill_h = PatternFill("solid", fgColor="1A1A1A")
    font_h = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_t = Font(name="Calibri", bold=True, size=16)
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    fill_a = PatternFill("solid", fgColor="F8D7DA")
    fill_b = PatternFill("solid", fgColor="FFF3CD")
    fill_c = PatternFill("solid", fgColor="F4F4F4")
    fill_iny = PatternFill("solid", fgColor="D6EAF8")

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(1, col)
            cell.fill = fill_h
            cell.font = font_h
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    def fill_clase(clase):
        return {"A": fill_a, "B": fill_b, "C": fill_c, "INY": fill_iny, "CUR": fill_iny}.get(clase, fill_c)

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Cruce FarmaCapital vs surtido Similares"
    ws["A1"].font = font_t
    ws.merge_cells("A1:B1")
    ws["A2"] = "Excel: pricing/fuentes/articulos_farmacias.xlsx · inventario vivo 20 ago 2026"
    ws["A3"] = "Cobertura = mismo genérico (principio + concentración + forma). No exige la misma marca comercial."
    rows_r = [
        ("Genéricos únicos Similares (medicamentos / curación / suplementos)", resumen["similares_unicos"]),
        ("Productos activos en tu inventario", resumen["inventario_activo"]),
        ("Ya cubiertos (los tienes, aunque sea otra marca)", resumen["cubiertos"]),
        ("  · con stock suficiente (hoy + pedido Levic 19-ago)", resumen["ok"]),
        ("  · hay que rellenar (bajo el mínimo sucursal)", resumen["rellenar"]),
        ("Huecos: Similares lo vende y tú no lo tienes", resumen["huecos"]),
        ("  · clase A (rotación alta) — comprar primero", resumen["huecos_a"]),
        ("  · clase B (rotación media)", resumen["huecos_b"]),
        ("  · clase C (especialidad)", resumen["huecos_c"]),
        ("  · inyectables", resumen["huecos_iny"]),
        ("Cobertura del surtido Similares", f"{resumen['cobertura_pct']}%"),
        ("Piezas a pedir — solo huecos clase A+B", resumen["pzas_huecos_ab"]),
        ("Piezas a pedir — todos los huecos", resumen["pzas_huecos"]),
        ("Piezas a pedir — rellenar lo que ya tienes", resumen["pzas_rellenar"]),
    ]
    ws["A5"] = "Métrica"
    ws["B5"] = "Valor"
    style_header(ws, 2)
    # overwrite row 5 as header after writing title — redo
    for i, (a, b) in enumerate(rows_r, start=6):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)

    ws["A22"] = "Política de stock (sucursal pequeña tipo Similares, no CEDIS)"
    ws["A22"].font = Font(name="Calibri", bold=True, size=12)
    headers_pol = ["Clase", "Qué es", "Mínimo", "Objetivo", "Si es hueco, pide"]
    for i, h in enumerate(headers_pol, 1):
        c = ws.cell(23, i, h)
        c.fill = fill_h
        c.font = font_h
    for i, row in enumerate([
        ("A", "Analgésicos, diabetes, presión, gastro, alergia, genéricos de mostrador", 6, 12, 6),
        ("B", "Antibióticos menos comunes, respiratorio, piel, vitaminas", 4, 8, 4),
        ("C", "Especialidad / baja rotación", 2, 4, 2),
        ("INY", "Inyectables (caducan y ocupan frío)", 2, 4, 2),
        ("CUR", "Material de curación de alta vuelta", 8, 12, 8),
    ], start=24):
        for j, v in enumerate(row, 1):
            ws.cell(i, j, v)

    ws["A31"] = (
        "No copies el almacén de una sucursal Similares (1,800–2,000 SKUs a tope). "
        "Una farmacia nueva cubre primero clase A, luego B. Clase C se pide de 2 en 2 cuando ya hay consulta. "
        "El pedido Levic del 19-ago ya se restó: si pediste 14 y tienes 1, el disponible es 15 y no vuelve a salir aquí. "
        "Compara vs Similares por genérico, no por marca: si tienes AMSA/Ultra del mismo PA, cuenta como cubierto."
    )
    ws.merge_cells("A31:E33")
    ws["A31"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 22

    def sheet_pedido(name, data, extra_fc=False):
        ws = wb.create_sheet(name)
        headers = [
            "Prioridad", "Clase", "Pedir", "SKU Similares", "Genérico (Similares)",
            "Concentración", "Presentación", "Línea", "Grupo",
            "Precio venta Similares", "Mín sucursal", "Objetivo",
        ]
        if extra_fc:
            headers += ["Tu SKU", "Tu producto", "Stock hoy", "Pedido Levic 19-ago", "Match %"]
        for i, h in enumerate(headers, 1):
            ws.cell(1, i, h)
        style_header(ws, len(headers))
        for r_i, r in enumerate(data, start=2):
            vals = [
                r_i - 1,
                r["clase"],
                r["pedir"],
                r["sim_sku"],
                r["descripcion"],
                r.get("concentracion") or "",
                r.get("contenido") or r.get("presentacion") or "",
                r.get("linea") or "",
                r.get("grupo") or "",
                r.get("precio") or 0,
                r["minimo"],
                r["objetivo"],
            ]
            if extra_fc:
                vals += [r.get("fc_sku"), r.get("fc_nombre"), r.get("stock"), r.get("incoming"), r.get("match")]
            for c_i, v in enumerate(vals, 1):
                cell = ws.cell(r_i, c_i, v)
                cell.border = thin
                if c_i == 2:
                    cell.fill = fill_clase(r["clase"])
            ws.cell(r_i, 3).font = Font(name="Calibri", bold=True, size=12)
        widths = [10, 8, 8, 14, 42, 16, 22, 22, 28, 16, 12, 12, 16, 36, 12, 16, 10]
        for i, w in enumerate(widths[:len(headers)], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(data)+1)}"
        return ws

    # Priority buy: A+B huecos first, then rellenar A
    p1 = [r for r in huecos if r["clase"] in ("A", "B", "CUR")]
    p_resto = [r for r in huecos if r["clase"] not in ("A", "B", "CUR")]
    sheet_pedido("1_Comprar_prioridad", p1)
    sheet_pedido("2_Huecos_especialidad", p_resto)
    sheet_pedido("3_Rellenar_lo_que_ya_tienes", rellenar, extra_fc=True)

    ws = wb.create_sheet("4_Ya_cubiertos")
    headers = ["Clase", "SKU Similares", "Genérico", "Tu SKU", "Tu producto", "Stock", "Levic 19-ago", "Match %"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    style_header(ws, len(headers))
    for r_i, r in enumerate(sorted(ok, key=lambda x: x["descripcion"]), start=2):
        for c_i, v in enumerate([
            r["clase"], r["sim_sku"], r["descripcion"], r["fc_sku"], r["fc_nombre"],
            r["stock"], r["incoming"], r["match"],
        ], 1):
            ws.cell(r_i, c_i, v).border = thin
    for i, w in enumerate([8, 14, 42, 16, 36, 10, 14, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
