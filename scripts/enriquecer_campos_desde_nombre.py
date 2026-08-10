#!/usr/bin/env python3
"""Genera SQL/CSV enriqueciendo marca, presentación y principio activo desde nombres."""

from __future__ import annotations

import csv
import hashlib
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from homologar_tickets_a_excel import sku_for_row  # noqa: E402
from parse_nombre_producto import parse_nombre_producto  # noqa: E402

DEFAULT_XLSX = Path(
    "/Users/ibarra/Library/CloudStorage/Dropbox/FarmaCapital/Tickets/"
    "FarmaCapital_inventario_homologado_completo.xlsx"
)
OUT_SQL = ROOT / "sql" / "actualizar_campos_producto_enriquecidos.sql"
OUT_CSV = ROOT / "sql" / "preview_campos_producto_enriquecidos.csv"


def sql_quote(val: str | None) -> str:
    if val is None or str(val).strip() == "":
        return "NULL"
    return "'" + str(val).replace("'", "''") + "'"


def sql_set(fields: dict[str, str | None]) -> str:
    parts = []
    for key, val in fields.items():
        parts.append(f"{key} = {sql_quote(val)}")
    return ", ".join(parts)


def main() -> None:
    if not DEFAULT_XLSX.exists():
        raise SystemExit(f"No existe Excel: {DEFAULT_XLSX}")

    wb = openpyxl.load_workbook(DEFAULT_XLSX, read_only=True, data_only=True)
    ws = wb["Compras_maestro"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    rows_out: list[dict[str, str]] = []
    stats = {"marca": 0, "presentacion": 0, "principio_activo": 0, "nombre_limpio": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["N.º ticket / orden"]]:
            continue
        ticket = str(row[idx["N.º ticket / orden"]])
        linea = str(row[idx["Línea ticket"]])
        nombre_orig = str(row[idx["Nombre / variante"]] or "").strip()
        tipo = row[idx["Tipo de producto"]]
        sku = sku_for_row(tuple(row))
        parsed = parse_nombre_producto(nombre_orig, tipo)

        fields = {
            "nombre": parsed.nombre if parsed.nombre != nombre_orig else None,
            "marca": parsed.marca,
            "presentacion": parsed.presentacion,
            "principio_activo": parsed.principio_activo,
            "concentracion": parsed.concentracion,
            "forma_farmaceutica": parsed.forma_farmaceutica,
        }
        fields = {k: v for k, v in fields.items() if v}
        if not fields:
            continue

        for k in ("marca", "presentacion", "principio_activo"):
            if fields.get(k):
                stats[k] += 1
        if fields.get("nombre"):
            stats["nombre_limpio"] += 1

        rows_out.append(
            {
                "sku": sku,
                "ticket": ticket,
                "linea": linea,
                "nombre_original": nombre_orig,
                "nombre_nuevo": fields.get("nombre", nombre_orig),
                "marca": fields.get("marca", ""),
                "presentacion": fields.get("presentacion", ""),
                "principio_activo": fields.get("principio_activo", ""),
                "concentracion": fields.get("concentracion", ""),
                "forma_farmaceutica": fields.get("forma_farmaceutica", ""),
            }
        )

    lines = [
        "-- Actualiza marca, presentación, principio activo y nombre limpio",
        f"-- Filas: {len(rows_out)}",
        "-- Ejecutar UNA vez en Supabase SQL Editor",
        "",
        "begin;",
        "",
    ]

    for item in rows_out:
        sku = item["sku"]
        update_fields = {}
        if item["nombre_nuevo"] and item["nombre_nuevo"] != item["nombre_original"]:
            update_fields["nombre"] = item["nombre_nuevo"]
        for col in ("marca", "presentacion", "principio_activo", "concentracion", "forma_farmaceutica"):
            if item[col]:
                update_fields[col] = item[col]
        if not update_fields:
            continue
        lines.append(f"-- {item['ticket']} L{item['linea']} | {item['nombre_original'][:70]}")
        lines.append(
            f"update public.productos set {sql_set(update_fields)} "
            f"where sku = {sql_quote(sku)};"
        )
        lines.append("")

    lines.extend(["commit;", ""])
    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")

    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()) if rows_out else [])
        if rows_out:
            w.writeheader()
            w.writerows(rows_out)

    print(f"SQL: {OUT_SQL}")
    print(f"Preview CSV: {OUT_CSV}")
    print(f"Filas: {len(rows_out)}")
    print("Stats:", stats)


if __name__ == "__main__":
    main()
