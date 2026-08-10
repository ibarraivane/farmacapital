#!/usr/bin/env python3
"""Genera SQL idempotente para completar productos/lotes faltantes tras carga parcial."""

from __future__ import annotations

import hashlib
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from homologar_tickets_a_excel import (  # noqa: E402
    MARGEN_VENTA,
    sql_block_for_row,
    sql_quote,
    sku_for_row,
)

DEFAULT_XLSX = Path(
    "/Users/ibarra/Library/CloudStorage/Dropbox/FarmaCapital/Tickets/"
    "FarmaCapital_inventario_homologado_completo.xlsx"
)
OUT = ROOT / "sql" / "carga_completar_faltantes_idempotente.sql"
MAX_BYTES = 900_000


def wrap_no_barcode_block(r: tuple, block: str, sku: str) -> str:
    m = re.search(
        r"select producto_id, lote_id from create_producto_with_lote\(([\s\S]*)\);\s*$",
        block.strip(),
    )
    if not m:
        return block
    inner = m.group(1)
    return f"""
-- idempotente {sku}
do $$
begin
  if not exists (select 1 from public.productos where sku = {sql_quote(sku)}) then
    perform producto_id, lote_id from create_producto_with_lote(
{inner}
    );
  end if;
end $$;
"""


def block_for_row_idempotent(r: tuple) -> str:
    bc = re.sub(r"\D", "", str(r[1] or ""))
    sku = sku_for_row(r)
    block = sql_block_for_row(r)
    if bc:
        m = re.search(
            r"select f\.producto_id, f\.lote_id into v_pid, v_lid\s+from create_producto_with_lote\(([\s\S]*?\))\s*f;",
            block,
        )
        if not m:
            return block
        create_args = m.group(1)
        label = block.split("\n")[1] if "\n" in block else f"-- {sku}"
        return f"""
{label}
do $$
declare v_pid bigint;
begin
  select id into v_pid from public.productos
  where sku = {sql_quote(sku)}
     or codigo_barras = {sql_quote(bc)}
  limit 1;
  if v_pid is null then
    perform producto_id, lote_id from create_producto_with_lote(
      {create_args}
    );
    insert into _fc_carga_map (codigo_barras, producto_id)
    select {sql_quote(bc)}, id from public.productos where codigo_barras = {sql_quote(bc)}
    on conflict (codigo_barras) do nothing;
  end if;
end $$;
"""
    return wrap_no_barcode_block(r, block, sku)


def main() -> None:
    if not DEFAULT_XLSX.exists():
        raise SystemExit(f"No existe: {DEFAULT_XLSX}")

    wb = openpyxl.load_workbook(DEFAULT_XLSX, read_only=True, data_only=True)
    ws = wb["Compras_maestro"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    rows: list[tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["N.º ticket / orden"]]:
            continue
        rows.append(tuple(row))

    header = f"""-- Completar carga idempotente (solo crea SKUs que falten)
-- Filas Excel: {len(rows)} | Margen: {int(MARGEN_VENTA * 100)}%
-- Uso: después de carga parcial _EJECUTAR_1..4 (ej. 433 productos, 972 pzas)
-- NO duplica productos existentes. Ejecutar UNA vez.

begin;

create temp table if not exists _fc_carga_map (
  codigo_barras text primary key,
  producto_id bigint
) on commit drop;

insert into _fc_carga_map (codigo_barras, producto_id)
select codigo_barras, id from public.productos
where codigo_barras is not null and btrim(codigo_barras) <> ''
on conflict (codigo_barras) do nothing;

"""

    footer = """
-- Resync stock desde lotes
update public.productos p
set stock = coalesce((
  select sum(l.cantidad_actual)
  from public.lotes l
  where l.producto_id = p.id and coalesce(l.activo, true) = true
), 0);

commit;

-- Verificar
select count(*) as productos_ticket
from public.productos where sku like 'FC-%' and sku not like 'FC100%';
select sum(cantidad_actual) as stock_lotes from public.lotes;
"""

    parts: list[str] = [header]
    current = header
    for r in rows:
        block = block_for_row_idempotent(r)
        if len((current + block).encode("utf-8")) > MAX_BYTES:
            parts.append(current)
            current = header
        current += block

    current += footer
    parts[0] = current if len(parts) == 1 else parts[0]
    if len(parts) > 1:
        parts[-1] = current

    if len(parts) == 1:
        OUT.write_text(current, encoding="utf-8")
        print(f"Generado: {OUT} ({len(rows)} filas)")
    else:
        for i, content in enumerate(parts, 1):
            p = ROOT / "sql" / f"carga_completar_faltantes_idempotente_{i}.sql"
            p.write_text(content, encoding="utf-8")
            print(f"Generado: {p}")
        print(f"Ejecutar _1 luego _2 ({len(rows)} filas total)")


if __name__ == "__main__":
    main()
