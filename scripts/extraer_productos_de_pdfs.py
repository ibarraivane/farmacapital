#!/usr/bin/env python3
"""
EXTRACTOR DE PRODUCTOS DESDE PDFS DE TICKETS
Procesa PDFs de compra farmacéutica y genera archivo Excel
con todos los productos, deduciendo información faltante.
"""

import json
import re
from pathlib import Path
from typing import Dict, List, Optional
import sys

# Mapa de deducción de abreviaturas comunes
ABREVIATURAS = {
    'CAPS': 'CAPSULAS',
    'CAP': 'CAPSULAS',
    'TAB': 'TABLETAS',
    'TABS': 'TABLETAS',
    'ML': 'MILILITROS',
    'MG': 'MILIGRAMOS',
    'G': 'GRAMOS',
    'GR': 'GRAMOS',
    'MCG': 'MICROGRAMOS',
    'SUSP': 'SUSPENSION',
    'SOL': 'SOLUCION',
    'INY': 'INYECTABLE',
    'COMP': 'COMPRIMIDOS',
    'GRAGEA': 'GRAGEAS',
    'POLVO': 'POLVO',
    'CREMA': 'CREMA',
    'POMADA': 'POMADA',
    'LOCION': 'LOCION',
    'GEL': 'GEL',
    'SPRAY': 'SPRAY',
    'GOTAS': 'GOTAS',
}

CATEGORIAS_AUTOMATICAS = {
    'AMOXICILINA': 'ANTIBIÓTICOS',
    'CEFALEXINA': 'ANTIBIÓTICOS',
    'IBUPROFENO': 'ANALGÉSICOS',
    'PARACETAMOL': 'ANALGÉSICOS',
    'DIPIRONA': 'ANALGÉSICOS',
    'DICLOFENACO': 'ANTIINFLAMATORIOS',
    'NIMESULIDA': 'ANTIINFLAMATORIOS',
    'OMEPRAZOL': 'DIGESTIVOS',
    'RANITIDINA': 'DIGESTIVOS',
    'LEVOTIROXINA': 'ENDOCRINOLOGÍA',
    'METFORMINA': 'ENDOCRINOLOGÍA',
    'VITAMINA': 'VITAMINAS',
    'LORATADINA': 'ANTIHISTAMÍNICOS',
    'CETIRIZINA': 'ANTIHISTAMÍNICOS',
    'CLOTRIMAZOL': 'DERMATOLOGÍA',
    'KETOCONAZOL': 'DERMATOLOGÍA',
}

class ProductoNormalizado:
    """Representa un producto extraído y normalizado"""

    def __init__(self, data: dict):
        self.codigo_barras = data.get('codigo_barras') or ''
        self.nombre = data.get('nombre', '').upper().strip()
        self.marca = data.get('marca', '').upper().strip()
        self.presentacion_raw = data.get('presentacion', '').strip()
        self.contenido_raw = data.get('contenido', '').strip()
        self.cantidad = self._parse_numero(data.get('cantidad'))
        self.precio_unitario = self._parse_precio(data.get('precio_unitario'))
        self.precio_total = self._parse_precio(data.get('precio_total'))
        self.fecha = data.get('fecha', '')
        self.proveedor = data.get('proveedor', 'DESCONOCIDO')

        # Normalizar
        self.presentacion_norm = self._normalizar_presentacion()
        self.contenido_norm, self.contenido_unidad = self._normalizar_contenido()
        self.categoria = self._deducir_categoria()
        self.unidad = self._deducir_unidad()

    @staticmethod
    def _parse_numero(valor) -> Optional[int]:
        """Parsea número de cantidad"""
        if not valor:
            return 1
        try:
            if isinstance(valor, (int, float)):
                return int(valor)
            return int(float(str(valor).replace(',', '.')))
        except:
            return 1

    @staticmethod
    def _parse_precio(valor) -> Optional[float]:
        """Parsea precio (pueden tener $ o comas)"""
        if not valor:
            return 0
        try:
            if isinstance(valor, (int, float)):
                return float(valor)
            # Eliminar $ y espacios
            clean = str(valor).replace('$', '').replace(',', '.').strip()
            return float(clean)
        except:
            return 0

    def _normalizar_presentacion(self) -> str:
        """Normaliza presentación: '40 CAP' → '40 CAPSULAS'"""
        if not self.presentacion_raw:
            return ''

        pres = self.presentacion_raw.upper().strip()

        # Reemplazar abreviaturas
        for abrev, completo in ABREVIATURAS.items():
            pres = re.sub(rf'\b{abrev}\b', completo, pres)

        return pres

    def _normalizar_contenido(self) -> tuple:
        """Normaliza contenido: '500MG' → (500, 'MG')"""
        if not self.contenido_raw:
            return None, None

        contenido = self.contenido_raw.upper().strip()

        # Buscar patrón: número + unidad
        match = re.match(r'(\d+(?:\.\d+)?)\s*([A-Z/]+)', contenido)
        if match:
            try:
                num = float(match.group(1))
                unidad = match.group(2).strip('/')
                return num, unidad
            except:
                return None, None

        return None, None

    def _deducir_categoria(self) -> str:
        """Deduce categoría según nombre del producto"""
        nombre_upper = self.nombre.upper()

        for keyword, categoria in CATEGORIAS_AUTOMATICAS.items():
            if keyword.upper() in nombre_upper:
                return categoria

        # Categoría por defecto según presentación
        if 'CREMA' in self.presentacion_norm or 'POMADA' in self.presentacion_norm:
            return 'DERMATOLOGÍA'

        return 'GENERAL'

    def _deducir_unidad(self) -> str:
        """Deduce unidad de medida desde presentación"""
        pres = self.presentacion_norm.upper()

        if 'CAPSULAS' in pres or 'CAPSULE' in pres:
            return 'CAPS'
        elif 'TABLETAS' in pres or 'COMPRIMIDOS' in pres:
            return 'TAB'
        elif 'ML' in pres or 'MILILITROS' in pres:
            return 'ML'
        elif 'GRAMOS' in pres or 'G' in pres:
            return 'G'
        elif 'AMPOLLETA' in pres or 'AMPOLLA' in pres:
            return 'AMP'
        elif 'SUSPENSION' in pres:
            return 'ML'
        elif 'SOLUCION' in pres:
            return 'ML'
        else:
            return 'UNIT'

    def to_dict(self) -> dict:
        """Convierte a diccionario para Excel"""
        return {
            'CODIGO_BARRAS': self.codigo_barras,
            'NOMBRE': self.nombre,
            'MARCA': self.marca,
            'PRESENTACION': self.presentacion_norm,
            'CONTENIDO': self.contenido_norm,
            'CONTENIDO_UNIDAD': self.contenido_unidad,
            'UNIDAD': self.unidad,
            'CANTIDAD': self.cantidad,
            'PRECIO_UNITARIO': round(self.precio_unitario, 2),
            'PRECIO_TOTAL': round(self.precio_total, 2),
            'CATEGORIA': self.categoria,
            'PROVEEDOR': self.proveedor,
            'FECHA': self.fecha,
        }


def extraer_productos_del_json(json_data: dict) -> List[ProductoNormalizado]:
    """Extrae productos de un JSON y los normaliza"""
    productos = []

    # Soporta múltiples formatos de respuesta de Claude
    productos_raw = json_data.get('productos', [])
    if not productos_raw:
        productos_raw = [json_data]  # Si es un solo producto

    for prod_data in productos_raw:
        try:
            prod = ProductoNormalizado(prod_data)
            if prod.nombre:  # Solo si tiene nombre
                productos.append(prod)
        except Exception as e:
            print(f"⚠️ Error procesando producto: {e}")

    return productos


def generar_sql_insert(productos: List[ProductoNormalizado], proveedor: str) -> str:
    """Genera SQL INSERT para cargar en Supabase"""

    sql_lines = [
        f"-- SQL INSERT: Productos de {proveedor}",
        f"-- Generado automáticamente",
        f"-- Total productos: {len(productos)}\n",
    ]

    for prod in productos:
        values = (
            f"'{prod.codigo_barras}'",
            f"'{prod.nombre}'",
            f"'{prod.marca}'",
            f"'{prod.presentacion_norm}'",
            f"'{prod.contenido_norm}'",
            f"'{prod.contenido_unidad}'",
            f"'{prod.unidad}'",
            f"{prod.precio_unitario}",
            f"'{prod.proveedor}'",
            f"'{prod.categoria}'",
            f"{prod.cantidad}",
            f"'{prod.fecha}'",
        )

        sql = f"""SELECT create_producto_con_oferta(
  jsonb_build_object(
    'codigo_barras', {values[0]},
    'nombre', {values[1]},
    'marca', {values[2]},
    'presentacion', {values[3]},
    'contenido', {values[4]},
    'contenido_unidad', {values[5]},
    'categoria', {values[9]},
    'tipo', 'MEDICAMENTO',
    'requiere_receta', false
  ),
  {values[10]},  -- cantidad
  {values[8]},   -- proveedor
  {values[7]},   -- precio_unitario
  {values[11]}::date,  -- fecha
  'LOTE-{prod.proveedor}-' || to_char(now(), 'YYYYMMDD'),
  {values[11]}::date + interval '1 year'  -- caducidad aprox
);"""

        sql_lines.append(sql)

    return '\n'.join(sql_lines)


def main():
    print("=" * 60)
    print("📦 EXTRACTOR DE PRODUCTOS DE PDFs")
    print("=" * 60)

    # Buscar JSONs generados por Claude Vision
    json_path = Path('/private/tmp/claude-501/-Users-ibarra-farmacapital/2bc1a13d-77ea-40e3-893c-378d4042f961/scratchpad/productos_extraidos.json')

    if not json_path.exists():
        print("\n❌ No se encontró archivo de productos extraídos")
        print("💡 Primero procesa los PDFs con Claude Vision usando:")
        print("   Tab 'Cargar PDF' en /admin/inventario")
        return

    # Leer JSONs
    with open(json_path) as f:
        datos = json.load(f)

    print(f"\n✓ Leyendo {len(datos)} proveedores...\n")

    todos_productos = []

    for proveedor, data in datos.items():
        if isinstance(data, str):
            print(f"⚠️ {proveedor}: datos sin procesar")
            continue

        productos = extraer_productos_del_json(data)
        todos_productos.extend(productos)

        print(f"✓ {proveedor}: {len(productos)} productos")

        # Mostrar ejemplos
        if productos:
            prod_ejemplo = productos[0]
            print(f"  Ejemplo: {prod_ejemplo.nombre}")
            print(f"  → {prod_ejemplo.presentacion_norm}, {prod_ejemplo.contenido_norm} {prod_ejemplo.contenido_unidad}")
            print(f"  → ${prod_ejemplo.precio_unitario} x {prod_ejemplo.cantidad} unidades\n")

    print(f"\n{'='*60}")
    print(f"📊 TOTAL: {len(todos_productos)} productos normalizados")
    print(f"{'='*60}\n")

    # Generar Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"

        # Encabezados
        headers = list(todos_productos[0].to_dict().keys()) if todos_productos else []
        ws.append(headers)

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for prod in todos_productos:
            ws.append(list(prod.to_dict().values()))

        # Ajustar ancho
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        # Guardar
        excel_path = Path('/Users/ibarra/farmacapital/PRODUCTOS_EXTRAIDOS.xlsx')
        wb.save(excel_path)
        print(f"✓ Excel generado: {excel_path}")

    except ImportError:
        print("⚠️ openpyxl no instalado, saltando Excel")
        print("   Instala: pip3 install openpyxl")

    # Generar SQL (mostrar muestra)
    print(f"\n📋 Mostrando SQL para los 3 primeros productos:\n")
    if todos_productos:
        sql = generar_sql_insert(todos_productos[:3], "MÚLTIPLES PROVEEDORES")
        print(sql)
        print(f"\n... ({len(todos_productos)-3} productos más)")

if __name__ == "__main__":
    main()
