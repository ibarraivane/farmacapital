#!/usr/bin/env python3
"""Genera SQL: campos parseados + precios con margen genérico 60% / patente 30%."""

from __future__ import annotations

import csv
import math
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from homologar_tickets_a_excel import sku_for_row  # noqa: E402
from parse_nombre_producto import COMMERCIAL_MED_NAMES, nombre_ticket_sucio, parse_nombre_producto  # noqa: E402

DEFAULT_XLSX = Path(
    "/Users/ibarra/Library/CloudStorage/Dropbox/FarmaCapital/Tickets/"
    "FarmaCapital_inventario_homologado_completo.xlsx"
)
OUT_SQL = ROOT / "sql" / "actualizar_catalogo_campos_y_precios.sql"
OUT_CSV = ROOT / "sql" / "preview_catalogo_campos_y_precios.csv"

MARGEN_GENERICO = 0.60
MARGEN_PATENTE = 0.30


def sql_quote(val: str | None) -> str:
    if val is None or str(val).strip() == "":
        return "NULL"
    return "'" + str(val).replace("'", "''") + "'"


def precio_venta(costo: float, margen: float) -> float:
    if costo <= 0:
        return 0.0
    return math.ceil(costo * (1 + margen) * 100) / 100


def clasificar_producto(tipo_excel: str | None, parsed, nombre: str) -> tuple[str, float]:
    """Retorna (tipo_db, margen). tipo_db: generico | marca."""
    tipo_u = str(tipo_excel or "").upper()
    first = (nombre or "").strip().split()[0].upper() if nombre else ""

    if tipo_u == "MEDICAMENTO":
        if parsed.principio_activo and not parsed.marca:
            return "generico", MARGEN_GENERICO
        if parsed.marca and parsed.marca.upper() in COMMERCIAL_MED_NAMES:
            return "marca", MARGEN_PATENTE
        if first in COMMERCIAL_MED_NAMES:
            return "marca", MARGEN_PATENTE
        if parsed.principio_activo:
            return "generico", MARGEN_GENERICO
        return "marca", MARGEN_PATENTE

    # OTC, higiene, abarrotes → margen 60%, tipo marca comercial
    return "marca", MARGEN_GENERICO


def main() -> None:
    if not DEFAULT_XLSX.exists():
        raise SystemExit(f"No existe Excel: {DEFAULT_XLSX}")

    wb = openpyxl.load_workbook(DEFAULT_XLSX, read_only=True, data_only=True)
    ws = wb["Compras_maestro"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    preview: list[dict] = []
    lines = [
        "-- Catálogo: nombre/marca/presentación/PA + precios",
        "-- Genéricos y no-medicamentos: +60% sobre costo",
        "-- Medicamento de patente / marca comercial: +30% sobre costo",
        "-- Ejecutar UNA vez. Luego recarga Inventario en Admin.",
        "",
        "begin;",
        "",
    ]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["N.º ticket / orden"]]:
            continue

        nombre_orig = str(row[idx["Nombre / variante"]] or "").strip()
        tipo_excel = row[idx["Tipo de producto"]]
        costo = float(row[idx["Costo unitario s/IVA"]] or 0)
        sku = sku_for_row(tuple(row))
        parsed = parse_nombre_producto(nombre_orig, tipo_excel)
        if parsed.notas_parser == "nombre_ticket_invalido":
            continue

        tipo_db, margen = clasificar_producto(tipo_excel, parsed, nombre_orig)
        precio = precio_venta(costo, margen)

        update: dict[str, str | float] = {}
        nombre_final = parsed.nombre or nombre_orig
        if parsed.nombre and not nombre_ticket_sucio(parsed.nombre):
            if nombre_ticket_sucio(nombre_orig) or parsed.nombre != nombre_orig:
                update["nombre"] = parsed.nombre
                nombre_final = parsed.nombre
        if parsed.marca:
            update["marca"] = parsed.marca
        if parsed.presentacion:
            update["presentacion"] = parsed.presentacion
        if parsed.principio_activo:
            update["principio_activo"] = parsed.principio_activo
        if parsed.concentracion:
            update["concentracion"] = parsed.concentracion
        if parsed.forma_farmaceutica:
            update["forma_farmaceutica"] = parsed.forma_farmaceutica
        if parsed.categoria:
            cat_map = {
                "Higiene bucal": "Higiene",
                "Higiene personal": "Higiene",
                "Higiene capilar": "Higiene",
                "Cuidado personal": "Cuidado personal",
                "Botiquín": "Botiquín",
                "Medicamento": "Otro",
                "Abarrotes": "Abarrotes",
            }
            update["categoria"] = cat_map.get(parsed.categoria, parsed.categoria)
        update["tipo"] = tipo_db
        if costo > 0:
            update["costo"] = round(costo, 2)
            update["precio"] = precio

        set_parts = []
        for k, v in update.items():
            if isinstance(v, (int, float)):
                set_parts.append(f"{k} = {v}")
            else:
                set_parts.append(f"{k} = {sql_quote(str(v))}")

        lines.append(f"-- {sku} | margen {int(margen*100)}% | {nombre_orig[:60]}")
        lines.append(
            f"update public.productos set {', '.join(set_parts)} where sku = {sql_quote(sku)};"
        )
        lines.append("")

        preview.append(
            {
                "sku": sku,
                "nombre_original": nombre_orig,
                "nombre": nombre_final,
                "marca": parsed.marca or "",
                "presentacion": parsed.presentacion or "",
                "principio_activo": parsed.principio_activo or "",
                "tipo": tipo_db,
                "margen_pct": int(margen * 100),
                "costo": costo,
                "precio_nuevo": precio,
            }
        )

    lines.extend(
        [
            "commit;",
            "",
            "-- Resumen",
            "select tipo, count(*) as n, round(avg(precio/costo - 1)*100,1) as margen_prom_pct",
            "from public.productos",
            "where sku like 'FC-%' and costo > 0",
            "group by tipo order by 1;",
            "",
        ]
    )

    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        if preview:
            w = csv.DictWriter(f, fieldnames=list(preview[0].keys()))
            w.writeheader()
            w.writerows(preview)

    print(f"SQL: {OUT_SQL} ({len(preview)} productos)")
    print(f"Preview: {OUT_CSV}")


if __name__ == "__main__":
    main()
