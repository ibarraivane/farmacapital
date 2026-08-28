#!/usr/bin/env python3
"""
Cruza el catálogo contra la lista de artículos de farmacia (Excel) para dos cosas:

  1. Proponer el `principio_activo` de los productos que no lo tienen. El Excel trae
     una columna MARCA COMERCIAL junto al genérico, así que resuelve directo lo que la
     API no puede: que "Wermy" es gabapentina y "Sibicos" es bifonazol.
  2. Sacar precios de referencia verificados, igual que sync_precios_similares.py pero
     con mejor materia prima: aquí la concentración, el contenido y la presentación
     vienen en columnas separadas en lugar de embutidos en el nombre del producto.

El verificador de match es el mismo criterio (identidad + concentración + cantidad +
forma) y la confianza que se guarda es la calculada.

Uso:
  python3 scripts/sync_precios_excel_farmacias.py --dry-run
  python3 scripts/sync_precios_excel_farmacias.py --apply
  python3 scripts/sync_precios_excel_farmacias.py --dry-run --excel otra_lista.xlsx
"""
from __future__ import annotations

import argparse
import csv
import importlib.util
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EXCEL_DEFAULT = ROOT / "pricing" / "fuentes" / "articulos_farmacias.xlsx"
REPORTE_DIR = ROOT / "pricing" / "reportes"

_spec = importlib.util.spec_from_file_location("sync_similares", Path(__file__).with_name("sync_precios_similares.py"))
sync = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(sync)

import requests  # noqa: E402
from rapidfuzz import fuzz, process  # noqa: E402

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala con: python3 -m pip install --user openpyxl")

# Umbral a partir del cual consideramos que es literalmente la misma marca comercial
MARCA_EXACTA = 92
MARCA_PARECIDA = 85


def cargar_excel(ruta: Path) -> list[dict]:
    """
    Columnas esperadas:
      0 SKU · 1 MARCA COMERCIAL · 2 DESCRIPCIÓN · 3 PRECIO CON IVA
      4 CONCENTRACION · 5 CONTENIDO DE LA PRESENTACIÓN · 7 PRESENTACION
    """
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: list[dict] = []
    for fila in ws.iter_rows(min_row=3, values_only=True):
        marca = (fila[1] or "").strip() if isinstance(fila[1], str) else ""
        desc = (fila[2] or "").strip() if isinstance(fila[2], str) else ""
        if not marca and not desc:
            continue
        try:
            precio = float(str(fila[3]).replace("$", "").replace(",", ""))
        except (TypeError, ValueError):
            continue
        if precio <= 0:
            continue
        concentracion = str(fila[4] or "")
        contenido = str(fila[5] or "")
        presentacion = str(fila[7] or "") if len(fila) > 7 else ""
        out.append({
            "marca": marca,
            "marca_norm": sync.normalizar(marca),
            "descripcion": desc,
            "desc_norm": sync.normalizar(desc),
            "precio": precio,
            "concentraciones": sync.extraer_concentraciones(f"{concentracion} {desc}"),
            "cantidad": sync.extraer_cantidad(contenido) or sync.extraer_cantidad(desc),
            "forma": sync.extraer_forma(presentacion) or sync.extraer_forma(desc),
            "etiqueta": " ".join(x for x in [marca, desc, concentracion, contenido] if x).strip(),
        })
    wb.close()
    return out


def tokens_poco_discriminativos(catalogo: list[dict], frac: float = 0.015) -> set[str]:
    """
    Palabras que aparecen en tantas filas del Excel que no identifican nada.
    Sin esto, un `principio_activo` como "Producto homeopático / natural" reduce el
    ingrediente a "natural" y cualquier fila que diga NATURAL cuenta como match.
    """
    from collections import Counter

    frecuencia: Counter[str] = Counter()
    for cand in catalogo:
        frecuencia.update(set(sync.tokens_utiles(cand["descripcion"])))
    minimo = max(40, int(len(catalogo) * frac))
    return {t for t, n in frecuencia.items() if n >= minimo}


def comparte_algo(att: dict, cand: dict) -> bool:
    """¿Hay al menos un término en común entre nuestro producto y el candidato?"""
    nuestros = set(sync.tokens_utiles(att["texto"]))
    suyos = set(sync.tokens_utiles(cand["descripcion"]))
    return bool(nuestros & suyos)


def nuestros_atributos(p: dict) -> dict:
    texto = " ".join(filter(None, [
        str(p.get("principio_activo") or ""),
        str(p.get("nombre") or ""),
        str(p.get("presentacion") or ""),
    ]))
    return {
        "texto": texto,
        "ingredientes": sync.ingredientes(p.get("principio_activo") or ""),
        "marca_norm": sync.normalizar(p.get("marca") or ""),
        "concentraciones": sync.extraer_concentraciones(texto),
        "cantidad": sync.extraer_cantidad(texto),
        "forma": sync.extraer_forma(texto) or sync.extraer_forma(p.get("forma_farmaceutica") or ""),
    }


def evaluar(att: dict, cand: dict, vagos: set[str]) -> tuple[int, list[str], bool]:
    """
    Devuelve (confianza 0-100, razones, marca_es_la_misma).

    La identidad se resuelve por marca comercial o por principio activo, lo que dé
    más señal; después se ajusta con los atributos que sí se pueden comparar.
    """
    razones: list[str] = []
    tope = 100

    # ── Identidad por marca comercial
    identidad_marca = 0
    marca_misma = False
    if att["marca_norm"] and cand["marca_norm"]:
        r = fuzz.ratio(att["marca_norm"], cand["marca_norm"])
        if r >= MARCA_EXACTA:
            identidad_marca, marca_misma = 90, True
            razones.append("misma marca comercial")
        elif r >= MARCA_PARECIDA and comparte_algo(att, cand):
            # Un parecido tipográfico solo no basta: el campo `marca` del catálogo a
            # veces trae al proveedor, y "Sumitex" (catéter) pegaba con "Sumitrex"
            # (sumatriptán). Se exige que además compartan algún término real.
            identidad_marca, marca_misma = 76, True
            razones.append(f"marca parecida ({r})")

    # ── Identidad por principio activo
    identidad_pa = 0
    if att["ingredientes"]:
        utiles = [i for i in att["ingredientes"] if i not in vagos]
        if not utiles:
            razones.append("principio activo demasiado generico")
        else:
            cand_tokens = sync.tokens_utiles(cand["descripcion"])
            presentes = [i for i in utiles if sync.token_presente(i, cand_tokens)]
            cobertura = len(presentes) / len(utiles)
            if cobertura == 1.0:
                identidad_pa = 85
                razones.append("principios activos completos")
            elif cobertura >= 0.6:
                identidad_pa = 66
                razones.append(f"{len(presentes)}/{len(utiles)} principios activos")
            elif cobertura > 0:
                identidad_pa = 35
                tope = min(tope, 60)
                razones.append(f"solo {len(presentes)}/{len(utiles)} principios activos")

    score = max(identidad_marca, identidad_pa)
    if score == 0:
        return 0, ["sin identidad comprobable"], False

    # ── Atributos comparables
    if att["concentraciones"] and cand["concentraciones"]:
        if att["concentraciones"] & cand["concentraciones"]:
            score += 10
            razones.append("concentracion coincide")
        else:
            score -= 30
            tope = min(tope, 68)
            razones.append(f"concentracion difiere {sorted(att['concentraciones'])}≠{sorted(cand['concentraciones'])}")
    elif cand["concentraciones"] and not att["concentraciones"]:
        # Varios SKUs nuestros pueden colapsar en la misma fila del Excel
        score -= 5
        tope = min(tope, 80)
        razones.append("nuestro catalogo sin concentracion")

    if att["cantidad"] and cand["cantidad"]:
        if att["cantidad"] == cand["cantidad"]:
            score += 8
            razones.append("cantidad coincide")
        else:
            score -= 25
            tope = min(tope, 76)
            razones.append(f"cantidad difiere {att['cantidad']}≠{cand['cantidad']}")

    if att["forma"] and cand["forma"]:
        if att["forma"] == cand["forma"]:
            score += 6
            razones.append("forma coincide")
        else:
            score -= 22
            tope = min(tope, 70)
            razones.append(f"forma difiere {att['forma']}≠{cand['forma']}")

    return max(0, min(tope, int(round(score)))), razones, marca_misma


def candidatos_para(att: dict, catalogo: list[dict], indice_marca: dict[str, list[int]],
                    desc_lista: list[str], marcas_lista: list[str],
                    indice_marca_nombre: dict[str, list[int]]) -> list[int]:
    """Preselecciona filas del Excel que valga la pena puntuar."""
    idx: set[int] = set()

    if att["marca_norm"]:
        idx.update(indice_marca.get(att["marca_norm"], []))
        for marca, score, pos in process.extract(
            att["marca_norm"], marcas_lista, scorer=fuzz.ratio, limit=3
        ):
            if score >= MARCA_PARECIDA:
                idx.update(indice_marca_nombre[marca])

    if att["ingredientes"]:
        consulta = " ".join(att["ingredientes"])
        for _, score, pos in process.extract(
            consulta, desc_lista, scorer=fuzz.token_set_ratio, limit=12
        ):
            if score >= 60:
                idx.add(pos)

    return list(idx)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default=str(EXCEL_DEFAULT))
    ap.add_argument("--fuente", default="similares")
    ap.add_argument("--umbral", type=int, default=sync.UMBRAL_DEFAULT)
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    ruta = Path(args.excel).expanduser()
    if not ruta.exists():
        sys.exit(f"No existe {ruta}")

    env = sync.cargar_env()
    url, key = env.get("REACT_APP_SUPABASE_URL", ""), env.get("REACT_APP_SUPABASE_ANON_KEY", "")
    if not url or not key:
        sys.exit("Faltan credenciales Supabase en .env")

    catalogo = cargar_excel(ruta)
    print(f"Filas utilizables del Excel: {len(catalogo)}")

    indice_marca: dict[str, list[int]] = {}
    indice_marca_nombre: dict[str, list[int]] = {}
    for i, c in enumerate(catalogo):
        indice_marca.setdefault(c["marca_norm"], []).append(i)
        indice_marca_nombre.setdefault(c["marca_norm"], []).append(i)
    marcas_lista = [m for m in indice_marca_nombre if m]
    desc_lista = [c["desc_norm"] for c in catalogo]
    vagos = tokens_poco_discriminativos(catalogo)
    print(f"Marcas comerciales distintas: {len(marcas_lista)}")
    print(f"Términos demasiado comunes para identificar: {len(vagos)}")

    productos = sync.fetch_productos(url, key)
    headers = {"apikey": key, "Authorization": f"Bearer {key}"}
    r = requests.get(
        f"{url}/rest/v1/producto_precios_referencia_actual",
        headers=headers,
        params={"select": "producto_id,confianza", "fuente": f"eq.{args.fuente}"},
        timeout=60,
    )
    r.raise_for_status()
    ya_cubiertos = {x["producto_id"] for x in r.json() if (x.get("confianza") or 0) >= args.umbral}
    print(f"Productos: {len(productos)} · ya con referencia verificada: {len(ya_cubiertos)}")

    precios: list[dict] = []
    propuestas_pa: list[dict] = []

    for p in productos:
        att = nuestros_atributos(p)
        pos = candidatos_para(att, catalogo, indice_marca, desc_lista, marcas_lista, indice_marca_nombre)
        if not pos:
            continue

        mejor = None
        for i in pos:
            cand = catalogo[i]
            conf, razones, marca_misma = evaluar(att, cand, vagos)
            if mejor is None or conf > mejor[0]:
                mejor = (conf, razones, marca_misma, cand)
        if mejor is None:
            continue
        conf, razones, marca_misma, cand = mejor

        # Propuesta de principio activo: solo cuando es literalmente la misma marca
        # comercial y nosotros no tenemos el dato. Requiere confirmación humana.
        if marca_misma and not att["ingredientes"]:
            propuestas_pa.append({
                "confirmado": "",
                "sku": p.get("sku"),
                "nombre_fc": p.get("nombre"),
                "marca_fc": p.get("marca"),
                "principio_activo_propuesto": cand["descripcion"],
                "marca_excel": cand["marca"],
                "precio_excel": cand["precio"],
            })

        if sync.descalificado(razones) or conf < args.umbral:
            continue
        costo = p.get("costo")
        costo = float(costo) if costo not in (None, "") else None
        if sync.validar_precio(cand["precio"], costo):
            continue
        if p["id"] in ya_cubiertos:
            continue

        precios.append({
            "producto_id": p["id"],
            "sku": p.get("sku"),
            "nombre": p.get("nombre"),
            "precio": cand["precio"],
            "confianza": conf,
            "nombre_fuente": cand["etiqueta"][:180],
            "razones": "; ".join(razones),
            "piezas_fuente": cand.get("cantidad"),
        })

    alta = [x for x in precios if x["confianza"] >= sync.UMBRAL_ALTA]
    print("\n─── Resultado ───")
    print(f"Precios nuevos verificados : {len(precios)}  (alta {len(alta)} · media {len(precios) - len(alta)})")
    print(f"Propuestas de principio activo: {len(propuestas_pa)}")

    REPORTE_DIR.mkdir(parents=True, exist_ok=True)
    hoy = date.today().isoformat()

    f_precios = REPORTE_DIR / f"excel_precios_{hoy}.csv"
    with f_precios.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["sku", "nombre", "precio", "confianza", "fila_excel", "razones"])
        for x in sorted(precios, key=lambda y: -y["confianza"]):
            w.writerow([x["sku"], x["nombre"], x["precio"], x["confianza"], x["nombre_fuente"], x["razones"]])
    print(f"CSV: {f_precios}")

    if propuestas_pa:
        f_pa = REPORTE_DIR / f"excel_propuestas_pa_{hoy}.csv"
        with f_pa.open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(propuestas_pa[0].keys()))
            w.writeheader()
            w.writerows(sorted(propuestas_pa, key=lambda x: (x["nombre_fc"] or "").lower()))
        print(f"CSV: {f_pa}")

    if args.dry_run or not args.apply:
        if not args.dry_run:
            print("Usa --apply para guardar los precios en Supabase")
        return
    if not precios:
        print("Nada que insertar.")
        return

    sync.tomar_lock()
    wh = {**headers, "Content-Type": "application/json", "Prefer": "return=representation"}
    imp = requests.post(
        f"{url}/rest/v1/importaciones_referencia",
        headers=wh,
        json={
            "fuente": args.fuente,
            "tipo": "venta",
            "fecha_lista": hoy,
            "archivo": ruta.name,
            "filas_ok": len(precios),
            "notas": f"sync_precios_excel_farmacias.py (match verificado, umbral {args.umbral})",
        },
        timeout=60,
    )
    imp.raise_for_status()
    import_id = imp.json()[0]["id"]
    payload = [
        {
            "producto_id": x["producto_id"],
            "fuente": args.fuente,
            "tipo": "venta",
            "precio": round(x["precio"], 2),
            "fecha": hoy,
            "nombre_fuente": x["nombre_fuente"],
            "confianza": x["confianza"],
            "origen": "import_csv",
            "import_id": import_id,
            "notas": (
                (f"piezas_fuente:{x['piezas_fuente']} | " if x.get("piezas_fuente") else "")
                + f"excel:{ruta.name} | {x['razones']}"
            )[:500],
        }
        for x in precios
    ]
    for i in range(0, len(payload), 100):
        r = requests.post(f"{url}/rest/v1/producto_precios_referencia", headers=wh,
                          json=payload[i: i + 100], timeout=120)
        r.raise_for_status()
    print(f"Guardadas {len(payload)} referencias (import_id={import_id}).")


if __name__ == "__main__":
    main()
