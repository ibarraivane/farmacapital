#!/usr/bin/env python3
"""
Pipeline semanal de pricing FarmaCapital.

Lee:
  - sql/preview_catalogo_campos_y_precios.csv   (tu catálogo: costo y precio de venta)
  - pricing/precios_proveedores/*.csv           (exports de Exprezo, Nadro, etc. -- formato Exprezo:
                                                  Categoria,Producto,Precio Mayoreo,Precio por Unidad)
  - sql/plantilla_precios_competencia.csv       (sku, precio_similares, precio_del_ahorro -- llenado a mano)

Genera:
  - pricing/Reporte_Semanal_Pricing.xlsx
      Hoja "Compra": mejor precio de compra encontrado por producto y de dónde
      Hoja "Venta sugerida": precio sugerido con base en Similares/Del Ahorro
"""
import glob
import re
import sys
from pathlib import Path

import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CATALOGO = ROOT / "sql" / "preview_catalogo_campos_y_precios.csv"
PROVEEDORES_DIR = ROOT / "pricing" / "precios_proveedores"
COMPETENCIA_CSV = ROOT / "sql" / "plantilla_precios_competencia.csv"
OUT_XLSX = ROOT / "pricing" / "Reporte_Semanal_Pricing.xlsx"

FONT = "Arial"


def norm(s):
    return re.sub(r"\s+", " ", str(s).lower().replace("-", " ").replace("/", " ")).strip()


def cargar_proveedores():
    frames = []
    for path in sorted(glob.glob(str(PROVEEDORES_DIR / "*.csv"))):
        try:
            df = pd.read_csv(path)
        except Exception as e:
            print(f"  (omitido {path}: {e})")
            continue
        if {"Producto", "Precio Mayoreo"}.issubset(df.columns):
            df["Precio Mayoreo"] = df["Precio Mayoreo"].astype(str).replace(r'[\$,]', '', regex=True)
            df["Precio Mayoreo"] = pd.to_numeric(df["Precio Mayoreo"], errors="coerce")
            df["proveedor"] = Path(path).stem.split("_")[0]
            df["prod_norm"] = df["Producto"].apply(norm)
            frames.append(df[["Producto", "Precio Mayoreo", "proveedor", "prod_norm"]])
        else:
            print(f"  (formato no reconocido, se ignora: {path})")
    if not frames:
        return pd.DataFrame(columns=["Producto", "Precio Mayoreo", "proveedor", "prod_norm"])
    return pd.concat(frames, ignore_index=True)


def mejor_precio_compra(marca, nombre, proveedores):
    if proveedores.empty:
        return None, None, None, 0
    marca = str(marca) if pd.notna(marca) else ""
    query = norm(f"{marca} {nombre}")
    pool = proveedores[proveedores["prod_norm"].str.contains(norm(marca), na=False)] if marca.strip() else proveedores
    if pool.empty:
        return None, None, None, 0
    best = process.extractOne(query, pool["prod_norm"].tolist(), scorer=fuzz.token_set_ratio)
    if best is None or best[1] < 70:
        return None, None, None, best[1] if best else 0
    _, score, idx = best
    row = pool.iloc[idx]
    return row["Producto"], row["Precio Mayoreo"], row["proveedor"], score


def main():
    df = pd.read_csv(CATALOGO)
    df["costo"] = pd.to_numeric(df["costo"], errors="coerce")
    df["precio_nuevo"] = pd.to_numeric(df["precio"], errors="coerce")
    df["nombre_original"] = df["nombre"]
    if "activo" in df.columns:
        df = df[df["activo"].astype(str).str.lower() != "false"]

    proveedores = cargar_proveedores()
    print(f"Catálogo: {len(df)} productos. Archivos de proveedores leídos: "
          f"{proveedores['proveedor'].nunique() if not proveedores.empty else 0}")

    competencia = None
    if COMPETENCIA_CSV.exists():
        competencia = pd.read_csv(COMPETENCIA_CSV, dtype={"sku": str})
        for col in ("precio_similares", "precio_del_ahorro"):
            if col in competencia.columns:
                competencia[col] = pd.to_numeric(competencia[col], errors="coerce")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Compra"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    normal_font = Font(name=FONT, size=10)
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["SKU", "Producto", "Marca", "Tu Costo Actual", "Mejor Proveedor Encontrado",
               "Producto Coincidente", "Precio Mayoreo", "Ahorro", "Confianza"]
    for i, h in enumerate(headers, start=1):
        c = ws1.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    row = 2
    for _, r in df.iterrows():
        prod, precio, prov, score = mejor_precio_compra(r.get("marca"), r["nombre_original"], proveedores)
        vals = [r["sku"], r["nombre_original"], r.get("marca", ""), round(float(r["costo"]), 2)]
        ws1.append(vals + [prov or "", prod or "", precio if precio else "",
                            round(float(r["costo"]) - precio, 2) if precio else "",
                            "Alta" if score >= 85 else ("Media" if score >= 70 else "")])
        if precio:
            fill = green if float(r["costo"]) > precio else red
            for col in range(5, 10):
                ws1.cell(row=row, column=col).fill = fill
        for col in range(1, 10):
            ws1.cell(row=row, column=col).font = normal_font
            ws1.cell(row=row, column=col).border = border
        row += 1

    widths1 = [14, 42, 16, 14, 18, 40, 14, 12, 10]
    for i, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Venta sugerida")
    headers2 = ["SKU", "Producto", "Tu Precio Actual", "Tu Costo", "Precio Similares",
                "Precio Del Ahorro", "Precio Sugerido", "Nota"]
    for i, h in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    if competencia is not None:
        merged = competencia.merge(df[["sku", "nombre_original", "precio_nuevo", "costo"]], on="sku", how="left")
        r = 2
        for _, row_ in merged.iterrows():
            sim = row_.get("precio_similares")
            aho = row_.get("precio_del_ahorro")
            precios_validos = [p for p in (sim, aho) if pd.notna(p)]
            costo = float(row_["costo"]) if pd.notna(row_["costo"]) else None
            sugerido, nota = "", ""
            if precios_validos:
                sugerido = round(sum(precios_validos) / len(precios_validos), 2)
                if costo and sugerido < costo * 1.15:
                    nota = "Ojo: cerca o debajo de tu costo + margen mínimo"
                    sugerido = round(costo * 1.15, 2)
            ws2.append([row_["sku"], row_.get("nombre_original", ""),
                        round(float(row_["precio_nuevo"]), 2) if pd.notna(row_.get("precio_nuevo")) else "",
                        round(costo, 2) if costo else "",
                        sim if pd.notna(sim) else "", aho if pd.notna(aho) else "",
                        sugerido, nota])
            r += 1
        for rr in range(2, r):
            for col in range(1, 9):
                ws2.cell(row=rr, column=col).font = normal_font
                ws2.cell(row=rr, column=col).border = border
    else:
        ws2.cell(row=2, column=1, value="Llena sql/plantilla_precios_competencia.csv y vuelve a correr este script.")

    widths2 = [14, 42, 14, 12, 16, 16, 16, 30]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Guardado: {OUT_XLSX}")


if __name__ == "__main__":
    main()
