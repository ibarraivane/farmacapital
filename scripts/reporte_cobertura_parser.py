#!/usr/bin/env python3
"""Reporte de cobertura del parser sobre los 627 productos del Excel."""

from __future__ import annotations

import csv
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from homologar_tickets_a_excel import sku_for_row  # noqa: E402
from parse_nombre_producto import parse_nombre_producto  # noqa: E402

XLSX = Path(
    "/Users/ibarra/Library/CloudStorage/Dropbox/FarmaCapital/Tickets/"
    "FarmaCapital_inventario_homologado_completo.xlsx"
)
OUT = ROOT / "sql" / "reporte_cobertura_parser.csv"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Compras_maestro"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    rows_out = []
    stats = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["N.º ticket / orden"]]:
            continue
        ticket = str(row[idx["N.º ticket / orden"]])
        nombre = str(row[idx["Nombre / variante"]] or "")
        tipo = row[idx["Tipo de producto"]]
        sku = sku_for_row(tuple(row))
        p = parse_nombre_producto(nombre, tipo)

        stats["total"] += 1
        if p.notas_parser:
            stats["invalido"] += 1
        if p.marca:
            stats["marca"] += 1
        if p.presentacion:
            stats["pres"] += 1
        if p.forma_farmaceutica:
            stats["forma"] += 1
        if p.principio_activo:
            stats["pa"] += 1

        rows_out.append(
            {
                "sku": sku,
                "ticket": ticket,
                "nombre_original": nombre,
                "nombre_nuevo": p.nombre,
                "marca": p.marca or "",
                "presentacion": p.presentacion or "",
                "forma": p.forma_farmaceutica or "",
                "principio_activo": p.principio_activo or "",
                "categoria": p.categoria or "",
                "notas": p.notas_parser or "",
            }
        )

    with OUT.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
        w.writeheader()
        w.writerows(rows_out)

    t = stats["total"]
    print(f"Reporte: {OUT}")
    print(f"Total: {t}")
    print(f"  Marca:         {stats['marca']}/{t} ({100*stats['marca']/t:.1f}%)")
    print(f"  Presentación: {stats['pres']}/{t} ({100*stats['pres']/t:.1f}%)")
    print(f"  Forma:         {stats['forma']}/{t} ({100*stats['forma']/t:.1f}%)")
    print(f"  PA:            {stats['pa']}/{t} ({100*stats['pa']/t:.1f}%)")
    print(f"  Inválidos OCR: {stats['invalido']}/{t}")


if __name__ == "__main__":
    main()
