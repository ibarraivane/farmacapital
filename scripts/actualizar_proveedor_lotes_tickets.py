#!/usr/bin/env python3
"""Asigna proveedor_id en lotes según tienda del ticket (Excel homologado)."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from homologar_tickets_a_excel import sku_for_row, sql_quote  # noqa: E402

XLSX = Path(
    "/Users/ibarra/Library/CloudStorage/Dropbox/FarmaCapital/Tickets/"
    "FarmaCapital_inventario_homologado_completo.xlsx"
)
OUT = ROOT / "sql" / "actualizar_proveedor_lotes_tickets.sql"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Compras_maestro"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    tiendas: set[str] = set()
    rows_map: list[tuple[str, str, str]] = []
    seen: set[tuple[str, str]] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["N.º ticket / orden"]]:
            continue
        ticket = str(row[idx["N.º ticket / orden"]])
        linea = str(row[idx["Línea ticket"]])
        tienda = str(row[idx["Proveedor / lugar de compra"]] or "").strip()
        if not tienda:
            continue
        tiendas.add(tienda)
        sku = sku_for_row(tuple(row))
        lote = str(row[idx["Lote"]] or f"TK-{ticket}-{linea}").strip()
        key = (sku, lote)
        if key in seen:
            continue
        seen.add(key)
        rows_map.append((sku, lote, tienda))

    lines = [
        "-- FarmaCapital — Proveedor del lote = tienda de compra (ticket)",
        f"-- Mapeos únicos: {len(rows_map)} | Tiendas: {len(tiendas)}",
        "-- Requiere antes: sql/patch_proveedor_tienda_en_lotes.sql",
        "",
        "begin;",
        "",
        "-- 1) Catálogo de tiendas",
    ]

    for tienda in sorted(tiendas):
        lines.append(
            f"select public.fc_resolver_proveedor_tienda({sql_quote(tienda)}::text);"
        )

    lines.extend(
        [
            "",
            "-- 2) Mapeo SKU + lote → tienda",
            "create temp table _fc_proveedor_lote (",
            "  sku text not null,",
            "  numero_lote text not null,",
            "  tienda text not null,",
            "  primary key (sku, numero_lote)",
            ") on commit drop;",
            "",
            "insert into _fc_proveedor_lote (sku, numero_lote, tienda) values",
        ]
    )

    value_lines = [
        f"  ({sql_quote(sku)}, {sql_quote(lote)}, {sql_quote(tienda)})"
        for sku, lote, tienda in rows_map
    ]
    lines.append(",\n".join(value_lines) + ";")

    lines.extend(
        [
            "",
            "-- 3) Actualización masiva",
            "update public.lotes l",
            "set proveedor_id = public.fc_resolver_proveedor_tienda(m.tienda)",
            "from public.productos p",
            "join _fc_proveedor_lote m on m.sku = p.sku",
            "where l.producto_id = p.id",
            "  and l.numero_lote = m.numero_lote",
            "  and coalesce(l.activo, true);",
            "",
            "-- 4) Verificación",
            "select",
            "  count(*) as lotes_activos,",
            "  count(*) filter (where l.proveedor_id is not null) as con_tienda,",
            "  count(*) filter (where l.proveedor_id is null) as sin_tienda",
            "from public.lotes l",
            "where coalesce(l.activo, true);",
            "",
            "select pv.nombre as tienda, count(*) as lotes",
            "from public.lotes l",
            "join public.proveedores pv on pv.id = l.proveedor_id",
            "where coalesce(l.activo, true)",
            "group by pv.nombre",
            "order by lotes desc;",
            "",
            "commit;",
        ]
    )

    OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"SQL: {OUT}")
    print(f"Tiendas: {len(tiendas)}")
    print(f"Mapeos: {len(rows_map)}")


if __name__ == "__main__":
    main()
