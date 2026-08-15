#!/usr/bin/env python3
"""
Homologa tickets PDF (OCR Vision macOS) a hoja Compras_maestro.
Genera Excel consolidado listo para revisión / importación a inventario.
"""

from __future__ import annotations

import hashlib
import math
import re
import subprocess
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook

TICKETS_DIR = Path("/Users/ibarra/Library/CloudStorage/Dropbox/FarmaCapital/Tickets")
SOURCE_XLSX = TICKETS_DIR / "FarmaCapital_extraccion_todos_los_PDFs_08-08-2026.xlsx"
OUTPUT_XLSX = TICKETS_DIR / "FarmaCapital_inventario_homologado_completo.xlsx"
OCR_CACHE_DIR = Path(__file__).resolve().parent.parent / ".tmp_ocr_vision"
SWIFT_OCR = Path(__file__).resolve().parent / "pdf_vision_ocr.swift"
SQL_DIR = Path(__file__).resolve().parent.parent / "sql" / "generated"
MARGEN_VENTA = 0.35
MAX_SQL_BYTES = 42 * 1024

PDF_FILES = (
    "Bodega F-42.pdf",
    "El surtidor de su farmacia.pdf",
    "Farma Mx.pdf",
    "FarmaLive.pdf",
    "IFC 1.pdf",
    "IFC 2.pdf",
)

HEADERS = (
    "Línea ticket",
    "Código de barras",
    "Tipo de producto",
    "Marca",
    "Nombre / variante",
    "Presentación",
    "Contenido",
    "Unidad",
    "Cantidad",
    "Costo unitario s/IVA",
    "Costo total línea s/IVA",
    "Caducidad",
    "Lote",
    "Proveedor / lugar de compra",
    "Ubicación proveedor",
    "Fecha compra",
    "N.º ticket / orden",
    "Descripción original ticket",
    "Estado captura",
    "Notas",
)


def money(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    s = str(val).replace("$", "").replace(",", ".").strip()
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def norm_barcode(raw: str | None) -> str | None:
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) >= 8:
        return digits
    return None


def infer_tipo(nombre: str) -> str:
    n = nombre.upper()
    if any(k in n for k in ("CAPS", "TAB", "GRAG", "JARABE", "SOL ", "SUSP", "INY", "MEDIC", "HIS")):
        return "Medicamento"
    if any(k in n for k in ("DESOD", "DEO ", "JBN", "JABON", "SH ", "CRA ", "AC ", "CHAMP")):
        return "Higiene personal"
    if any(k in n for k in ("JERINGA", "CATETER", "VENDA", "CINTA", "PROTEC", "GUANTE")):
        return "Material médico"
    return "Producto"


def row_tuple(**kw) -> tuple:
    return tuple(kw.get(h, None) for h in HEADERS)


def ocr_lines(ocr: str) -> list[str]:
    return [
        l.strip()
        for l in ocr.splitlines()
        if l.strip() and not l.startswith("--- page")
    ]


def load_ocr_from_pdfs(force: bool = False) -> dict[str, str]:
    OCR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    ocr: dict[str, str] = {}
    for pdf_name in PDF_FILES:
        pdf_path = TICKETS_DIR / pdf_name
        if not pdf_path.exists():
            raise SystemExit(f"No existe PDF: {pdf_path}")
        cache_path = OCR_CACHE_DIR / f"{pdf_name.replace('.pdf', '')}.txt"
        if (
            force
            or not cache_path.exists()
            or pdf_path.stat().st_mtime > cache_path.stat().st_mtime
        ):
            result = subprocess.run(
                ["swift", str(SWIFT_OCR), str(pdf_path)],
                capture_output=True,
                text=True,
                check=False,
            )
            if result.returncode != 0:
                raise SystemExit(
                    f"OCR falló para {pdf_name}: {result.stderr[:400]}"
                )
            cache_path.write_text(result.stdout, encoding="utf-8")
        ocr[pdf_name] = cache_path.read_text(encoding="utf-8")
    return ocr


def load_existing_maestro(source: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb["Compras_maestro"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return [tuple(r) for r in rows if r and r[16] in ("77827", "440393")]


def find_money_values(text: str, limit: float = 5000.0) -> list[float]:
    vals: list[float] = []
    for m in re.finditer(r"\$\s*([\d ]+[.,][\d ]{2})", text):
        vals.append(money(m.group(1)))
    for m in re.finditer(r"(?<!\d)([\d]{1,4}[.,][\d]{2})(?!\d)", text):
        v = money(m.group(1))
        if 0 < v < limit:
            vals.append(v)
    return vals


def fix_fmx_qty(raw: str, line: str) -> int:
    s = raw.lstrip("-").replace(",", ".")
    if "." in s:
        return max(1, round(float(s)))
    v = int(float(s))
    if re.search(r"\b100\s+pza\s+950\b", line, re.I):
        return 1
    if v >= 100 and v % 100 == 0 and v <= 9900:
        candidate = v / 100
        if candidate == int(candidate) and 1 <= candidate <= 99:
            return int(candidate)
    if 1 <= v <= 99:
        return v
    return 1


def fmx_subtotal_from_block(block: list[str]) -> float:
    for bl in block:
        m = re.search(r"\$\s*([\d ]+)\s+([\d]{2})\b", bl)
        if m:
            v = money(f"{m.group(1).replace(' ', '')}.{m.group(2)}")
            if 0 < v < 5000:
                return v
        m = re.search(r"\$\s*([\d]+(?:\.[\d]{2})?)", bl)
        if m:
            v = money(m.group(1))
            if 0 < v < 5000:
                return v
    for bl in block:
        if "IVATRA" in bl or bl.startswith("10010"):
            continue
        m = re.search(r"(\d{2,4})\s+(\d{2})\b", bl)
        if m:
            v = money(f"{m.group(1)}.{m.group(2)}")
            if 1 <= v <= 3000:
                return v
    decimals: list[float] = []
    for bl in block:
        if "IVATRA" in bl or bl.startswith("10010"):
            break
        for m in re.finditer(r"(?<!\d)(\d{1,4}\.\d{2})(?!\d)", bl):
            v = float(m.group(1))
            if 0 < v < 5000:
                decimals.append(v)
    if decimals:
        plausible = [d for d in decimals if d >= 3]
        return plausible[-1] if plausible else decimals[-1]
    return 0.0


def parse_bodega(ocr: str) -> list[tuple]:
    proveedor = "Bodega F-42 Ejidos del Moral"
    ubic = "Iztapalapa, CDMX"
    fecha = "2026-08-08"
    ticket = "77827"
    text = ocr.replace("7509952933307", "7509552933307")
    items: dict[int, dict] = {}
    last_line = 0
    pattern = re.compile(
        r"(?:^|\n)(\d{1,3})\.\s*-?\s*([\d/]{8,16})|(?:^|\n)\.?(750\d{10,13})",
        re.M,
    )
    for m in pattern.finditer(text):
        if m.group(1):
            line_no = int(m.group(1))
            bc = norm_barcode(m.group(2))
        else:
            last_line += 1
            line_no = last_line
            bc = norm_barcode(m.group(3))
        if not bc:
            continue
        last_line = line_no
        if line_no > 250:
            continue
        tail = text[m.end() : m.end() + 220]
        qm = re.search(r"\([xX](\d+)\)", tail)
        qty = int(qm.group(1)) if qm else 1
        desc_m = re.search(r"\([xX]\d+\)\s*([^\n\r|]+)", tail)
        desc = (desc_m.group(1).strip() if desc_m else "").strip(" .=")
        prices = find_money_values(tail, 500.0)
        subtotal = prices[-1] if prices else 0.0
        if subtotal <= 0 and prices:
            subtotal = prices[0] * qty
        items[line_no] = {
            "line": line_no,
            "barcode": bc,
            "qty": qty,
            "subtotal": subtotal,
            "desc": desc,
        }
    rows = []
    for line_no in sorted(items):
        it = items[line_no]
        nombre = it["desc"] or f"Producto ticket {line_no}"
        qty = it["qty"]
        subtotal = it["subtotal"]
        rows.append(
            row_tuple(
                **{
                    "Línea ticket": line_no,
                    "Código de barras": it["barcode"],
                    "Tipo de producto": infer_tipo(nombre),
                    "Nombre / variante": nombre.title()[:120],
                    "Cantidad": qty,
                    "Costo unitario s/IVA": round(subtotal / max(qty, 1), 2),
                    "Costo total línea s/IVA": subtotal,
                    "Proveedor / lugar de compra": proveedor,
                    "Ubicación proveedor": ubic,
                    "Fecha compra": fecha,
                    "N.º ticket / orden": ticket,
                    "Descripción original ticket": it["desc"] or nombre,
                    "Estado captura": "Vision OCR PDF",
                    "Notas": "Bodega F-42.pdf",
                }
            )
        )
    return rows


def parse_surtidor(ocr: str) -> list[tuple]:
    proveedor = "El Surtidor de su Farmacia"
    ticket = "112558"
    fecha = "2026-08-08"
    ubic = "Central de Abasto, Iztapalapa, CDMX"
    lines = ocr_lines(ocr)
    bc_re = re.compile(r"(750\d{10}|759\d{10}|331\d{10}|650\d{10})")
    rows = []
    seen: set[str] = set()
    line_no = 0
    for i, line in enumerate(lines):
        bc_match = bc_re.search(line.replace(" ", ""))
        if not bc_match:
            continue
        bc = bc_match.group(1)
        if bc in seen:
            continue
        seen.add(bc)
        qty = 1
        name = ""
        for j in range(max(0, i - 5), i):
            cand = lines[j].strip()
            if re.fullmatch(r"\d+", cand) and int(cand) < 500:
                qty = int(cand)
            elif (
                len(cand) > 4
                and not bc_re.search(cand.replace(" ", ""))
                and "IVA" not in cand
                and re.search(r"[A-ZÁÉÍÓÚÑa-záéíóúñ]", cand)
            ):
                clean = re.sub(r"^[\*\-]+", "", cand).strip()
                if len(clean) > len(name):
                    name = clean.upper()
        nums: list[float] = []
        for j in range(i + 1, min(i + 7, len(lines))):
            if bc_re.search(lines[j].replace(" ", "")):
                break
            nums.extend(find_money_values(lines[j], 5000.0))
        nums = [n for n in nums if n > 0]
        subtotal = nums[-1] if nums else 0.0
        if subtotal <= 0 and nums:
            subtotal = nums[0] * qty
        line_no += 1
        rows.append(
            row_tuple(
                **{
                    "Línea ticket": line_no,
                    "Código de barras": bc,
                    "Tipo de producto": infer_tipo(name),
                    "Nombre / variante": (name or f"Producto {bc}")[:120],
                    "Cantidad": qty,
                    "Costo unitario s/IVA": round(subtotal / max(qty, 1), 2) if subtotal else 0,
                    "Costo total línea s/IVA": subtotal,
                    "Proveedor / lugar de compra": proveedor,
                    "Ubicación proveedor": ubic,
                    "Fecha compra": fecha,
                    "N.º ticket / orden": ticket,
                    "Descripción original ticket": name or bc,
                    "Estado captura": "Vision OCR PDF",
                    "Notas": "El surtidor de su farmacia.pdf",
                }
            )
        )
    return rows


def parse_ifc(ocr: str, ticket: str, folio: str) -> list[tuple]:
    proveedor = "IFC F8 Tienda"
    fecha = "2026-08-08"
    ubic = "Contreras, CDMX"
    text = (
        ocr.replace("PIEzA", "PIEZA")
        .replace("PAG", "PAQ")
        .replace("r16Za", "PIEZA")
        .replace("B.00", "3.00")
    )
    lines = ocr_lines(text)
    skip = (
        "IFC F8",
        "SAN BERN",
        "TEL:",
        "RFC",
        "REGIMEN",
        "CLIENTE",
        "ATENDIDO",
        "CANT P",
        "MAYOREO",
        "MENUDEO",
        "TOTAL",
        "EXPEDIDO",
        "FACTURA",
        "IZTAPALAPA",
        "Quinientos",
        "Un Mil",
        "Trescientos",
        "Articulos",
        "Productos",
    )
    rows = []
    line_no = 0
    for i, line in enumerate(lines):
        m = re.search(r"^(\d+\.?\d*)\s*(PIEZA|PZ|PAQ)", line, re.I)
        if not m:
            continue
        qty = max(1, round(float(m.group(1).replace(",", "."))))
        if qty > 200:
            continue
        monies: list[float] = []
        for j in range(i + 1, min(i + 4, len(lines))):
            bl = lines[j]
            if re.search(r"PIEZA|PZ|PAQ", bl, re.I):
                break
            if re.fullmatch(r"\d+\.\d{2}", bl):
                v = money(bl)
                if 0 < v < 500:
                    monies.append(v)
            if len(monies) >= 2:
                break
        if not monies:
            continue
        unit = monies[0]
        subtotal = monies[-1]
        if subtotal < unit and qty > 1:
            subtotal = round(unit * qty, 2)
        name_parts: list[str] = []
        for j in range(max(0, i - 4), i):
            cand = lines[j]
            if any(cand.startswith(s) for s in skip):
                continue
            if re.fullmatch(r"\d{4,8}", cand):
                continue
            if re.search(r"^[A-ZÁÉÍÓÚÑ/]", cand) and len(cand) > 4:
                name_parts.append(cand)
        name = " ".join(name_parts[-2:]).strip()
        line_no += 1
        rows.append(
            row_tuple(
                **{
                    "Línea ticket": line_no,
                    "Tipo de producto": infer_tipo(name),
                    "Nombre / variante": (name or f"Producto IFC {line_no}")[:120],
                    "Cantidad": qty,
                    "Costo unitario s/IVA": unit,
                    "Costo total línea s/IVA": subtotal,
                    "Proveedor / lugar de compra": proveedor,
                    "Ubicación proveedor": ubic,
                    "Fecha compra": fecha,
                    "N.º ticket / orden": ticket,
                    "Descripción original ticket": name,
                    "Estado captura": "Vision OCR PDF",
                    "Notas": f"IFC folio {folio}",
                }
            )
        )
    return rows


def parse_farma_mx(ocr: str) -> list[tuple]:
    proveedor = "Farma MX"
    ticket = "FMX-080826"
    fecha = "2026-08-08"
    ubic = "Iztapalapa, CDMX"
    lines = ocr_lines(ocr)
    rows = []
    line_no = 0
    i = 0
    while i < len(lines):
        if not re.fullmatch(r"\d{5,7}", lines[i]):
            i += 1
            continue
        clave = lines[i]
        qty = 1
        subtotal = 0.0
        name = ""
        lote = None
        cad = None
        block: list[str] = []
        j = i + 1
        while j < len(lines) and j < i + 14:
            bl = lines[j]
            if re.fullmatch(r"\d{5,7}", bl):
                break
            block.append(bl)
            if bl.startswith("L "):
                lm = re.search(r"FC\s+(\d{2}/\d{2}/\d{4})", bl, re.I)
                if lm:
                    d = lm.group(1).split("/")
                    cad = f"{d[2]}-{d[1]}-{d[0]}"
                lote = bl.replace("L ", "").split(" FC")[0].strip()
            elif (
                len(bl) > 8
                and not bl.startswith("10010")
                and not bl.startswith("IVATRA")
                and not re.fullmatch(r"[\d.]+", bl)
                and not re.search(r"pz[aei]?\b", bl, re.I)
                and "SUBTOTAL" not in bl.upper()
                and "FECHA DE CADUCIDAD" not in bl.upper()
            ):
                if re.search(r"[A-ZÁÉÍÓÚ\-]", bl):
                    name = bl
            j += 1
        for k, bl in enumerate(block):
            if not re.search(r"pz[aei]?\b", bl, re.I):
                continue
            qm = re.search(r"(-?\d+\.?\d*)\s*pz[aei]?\b", bl, re.I)
            if qm:
                qty = fix_fmx_qty(qm.group(1), bl)
            elif k > 0 and re.fullmatch(r"\d+\.\d{2}", block[k - 1]):
                qty = fix_fmx_qty(block[k - 1], bl)
            break
        qty = min(qty, 50)
        subtotal = fmx_subtotal_from_block(block)
        if subtotal <= 0:
            i += 1
            continue
        unit = round(subtotal / max(qty, 1), 2)
        line_no += 1
        rows.append(
            row_tuple(
                **{
                    "Línea ticket": line_no,
                    "Tipo de producto": "Medicamento",
                    "Nombre / variante": (name or f"Clave {clave}")[:120],
                    "Cantidad": qty,
                    "Costo unitario s/IVA": unit,
                    "Costo total línea s/IVA": subtotal,
                    "Caducidad": cad,
                    "Lote": lote,
                    "Proveedor / lugar de compra": proveedor,
                    "Ubicación proveedor": ubic,
                    "Fecha compra": fecha,
                    "N.º ticket / orden": ticket,
                    "Descripción original ticket": name or clave,
                    "Estado captura": "Vision OCR PDF",
                    "Notas": f"Farma MX clave {clave}",
                }
            )
        )
        i = j if j > i else i + 1
    return rows


def farmalive_barcode(line: str) -> str | None:
    """Extrae EAN de línea OCR FarmaLive (750…, 354…, 65024… Genomma, 366…, 780…)."""
    if not line or not line.strip():
        return None

    patterns = [
        r"[\[(]?\s*(750\d{10,11})",
        r"[\[(]?(354\d{10,11})",
        r"[\[(]?\s*(840\d{10,11})",
        r"[\[(]?\s*(780\d{10,11})",
        r"[\[(]?\s*(366\d{11,12})",
        r"[\[(]?\s*(65024\d{7,9})",
        r"(750\d{10,11})",
        r"(354\d{10,11})",
        r"(65024\d{7,9})",
        r"(780\d{10,11})",
        r"(366\d{11,12})",
    ]
    for p in patterns:
        m = re.search(p, line, re.I)
        if m:
            bc = norm_barcode(m.group(1))
            if bc and len(bc) >= 10:
                return bc

    compact = re.sub(r"\D", "", line)
    for run in re.findall(r"\d{12,14}", compact):
        if run.startswith(("750", "354", "65024", "366", "780")):
            return run

    for bare, prefixed in (
        ("2250343072", "7502250343072"),
        ("222503430721", "7502250343072"),
        ("1354312225027", "3543122250276"),
        ("543122250227", "3543122250276"),
        ("022503405381", "750022503405381"),
        ("1312250181", "7501312250181"),
    ):
        if bare in compact:
            return prefixed

    return None


def parse_farmalive(ocr: str) -> list[tuple]:
    proveedor = "FarmaLive"
    ticket = "FL-080826"
    fecha = "2026-08-08"
    ubic = "Chinampac de Juárez, Iztapalapa, CDMX"
    lines = ocr_lines(ocr)
    rows = []
    seen: set[str] = set()
    line_no = 0
    i = 0
    while i < len(lines):
        bc = farmalive_barcode(lines[i])
        if not bc or bc in seen:
            i += 1
            continue
        block = lines[i : min(i + 18, len(lines))]
        name_parts: list[str] = []
        qty = 1
        subtotal = 0.0
        for bl in block[1:]:
            if farmalive_barcode(bl):
                break
            if re.search(r"^Descto:", bl, re.I):
                dm = re.search(
                    r"Descto:.*?[\$#฿]\s*([\d,]+\.?\d*).*?[\$#฿]\s*([\d,]+\.?\d*)",
                    bl,
                    re.I,
                )
                if dm:
                    subtotal = money(dm.group(2))
                continue
            qm = re.search(r"^(\d+)\s+PACK\b", bl, re.I)
            if qm:
                qty = max(qty, int(qm.group(1)))
            qm2 = re.search(r"^(\d+)\s*[\$#฿]", bl)
            if qm2 and int(qm2.group(1)) < 500:
                qty = max(qty, int(qm2.group(1)))
            if re.search(r"[\$#฿]\s*([\d,]+\.?\d*)\s*$", bl):
                v = money(re.search(r"[\$#฿]\s*([\d,]+\.?\d*)\s*$", bl).group(1))
                if 0 < v < 5000:
                    subtotal = v if subtotal <= 0 else subtotal
            if (
                len(bl) > 3
                and not bl.startswith("Descto")
                and not re.fullmatch(r"[\d.]+", bl)
                and "OTC" not in bl
                and not re.search(r"^[\$#฿]", bl)
                and not re.search(r"^(RB|BAYER|GENOMMA|KSK|HEALTH|LAB)\b", bl)
            ):
                if re.search(r"[A-ZÁÉÍÓÚa-záéíóú]", bl):
                    name_parts.append(bl)
        tail = block[0] + " " + " ".join(block[1:6])
        m_tail = re.search(r"[\[(]?(750\d{10,11}|354\d{10,11})[\])]?\s*(.+)", tail)
        if m_tail and m_tail.group(2).strip():
            name_parts.insert(0, m_tail.group(2).strip())
        name = re.sub(r"\s+", " ", " ".join(name_parts[:4])).strip(" .|")
        if subtotal <= 0:
            money_vals = find_money_values(" ".join(block), 5000.0)
            subtotal = money_vals[-1] if money_vals else 0.0
        if subtotal <= 0:
            i += 1
            continue
        seen.add(bc)
        line_no += 1
        rows.append(
            row_tuple(
                **{
                    "Línea ticket": line_no,
                    "Código de barras": bc,
                    "Tipo de producto": infer_tipo(name),
                    "Nombre / variante": name.title()[:120] if name else f"Producto {bc}",
                    "Cantidad": qty,
                    "Costo unitario s/IVA": round(subtotal / max(qty, 1), 2),
                    "Costo total línea s/IVA": subtotal,
                    "Proveedor / lugar de compra": proveedor,
                    "Ubicación proveedor": ubic,
                    "Fecha compra": fecha,
                    "N.º ticket / orden": ticket,
                    "Descripción original ticket": name or bc,
                    "Estado captura": "Vision OCR PDF",
                    "Notas": "FarmaLive.pdf",
                }
            )
        )
        i += 1
    return rows


def write_workbook(rows: list[tuple], stats: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras_maestro"
    ws.append(HEADERS)
    for r in rows:
        ws.append(list(r))

    ws2 = wb.create_sheet("Resumen_homologacion")
    ws2.append(["Ticket", "Proveedor", "Líneas", "Piezas", "Costo s/IVA", "Objetivo ticket", "Notas"])
    targets = {
        "77827": "184 líneas / 206 pzas / $7,271 s/IVA",
        "440393": "162 líneas / 308 pzas / $10,548 s/IVA",
        "112558": "34 líneas / $5,461 total",
        "FMX-080826": "233 artículos / $5,310 s/IVA",
        "FL-080826": "20 artículos / 467 pzas / $24,586",
        "IFC1-080826": "19 prod / 60 pzas / $536",
        "IFC2-080826": "25 prod / 53 pzas / $1,331",
    }
    for s in stats:
        ws2.append(
            [
                s["ticket"],
                s["prov"],
                s["lines"],
                s["pieces"],
                s["cost"],
                targets.get(s["ticket"], ""),
                s["notes"],
            ]
        )

    ws3 = wb.create_sheet("Inventario_consolidado")
    cons_headers = (
        "Código de barras",
        "Tipo",
        "Marca",
        "Producto / variante",
        "Cantidad total",
        "Costo unitario prom",
        "Costo total",
        "Caducidad",
        "Lote",
        "Proveedor",
        "Ticket",
    )
    ws3.append(cons_headers)
    agg: dict[str, dict] = {}
    for r in rows:
        key = str(r[1] or f"{r[16]}:{r[4]}")
        a = agg.setdefault(
            key,
            {
                "bc": r[1],
                "tipo": r[2],
                "marca": r[3],
                "nombre": r[4],
                "qty": 0,
                "cost": 0.0,
                "cad": r[11],
                "lote": r[12],
                "prov": r[13],
                "ticket": r[16],
            },
        )
        a["qty"] += int(r[8] or 0)
        a["cost"] += float(r[10] or 0)
    for a in agg.values():
        qty = a["qty"] or 1
        ws3.append(
            [
                a["bc"],
                a["tipo"],
                a["marca"],
                a["nombre"],
                a["qty"],
                round(a["cost"] / qty, 2),
                round(a["cost"], 2),
                a["cad"],
                a["lote"],
                a["prov"],
                a["ticket"],
            ]
        )

    wb.save(OUTPUT_XLSX)


def sql_quote(val: Any) -> str:
    return "'" + str(val or "").replace("'", "''") + "'"


def sql_text_or_null(val: Any, max_len: int = 500) -> str:
    s = str(val or "").strip()
    if not s:
        return "NULL"
    return sql_quote(s[:max_len])


def sql_num_or_null(val: Any) -> str:
    if val is None or val == "":
        return "NULL"
    try:
        return str(round(float(val), 2))
    except (TypeError, ValueError):
        return "NULL"


def precio_venta(costo: float) -> float:
    if costo <= 0:
        return 0.0
    return math.ceil(costo * (1 + MARGEN_VENTA) * 100) / 100


def sku_for_row(r: tuple) -> str:
    bc = re.sub(r"\D", "", str(r[1] or ""))
    if len(bc) >= 8:
        return f"FC-{bc[-8:]}"
    digest = hashlib.md5(f"{r[16]}|{r[4]}|{r[0]}".encode()).hexdigest()[:8].upper()
    return f"FC-{digest}"


def sql_block_for_row(r: tuple) -> str:
    bc = re.sub(r"\D", "", str(r[1] or ""))
    sku = sku_for_row(r)
    costo = float(r[9] or 0)
    precio = precio_venta(costo)
    qty = max(1, int(r[8] or 1))
    lote = str(r[12] or f"TK-{r[16]}-{r[0]}").strip()
    cad = str(r[11] or "")[:10] if r[11] else None
    cad_sql = sql_quote(cad) if cad else "NULL"
    proveedor = sql_text_or_null(r[13])
    proveedor_arg = proveedor if proveedor != "NULL" else "NULL"
    tipo = "MEDICAMENTO" if r[2] == "Medicamento" else "GENERICO"
    desc = sql_quote(
        " — ".join(
            x
            for x in [str(r[4] or ""), str(r[5] or ""), f"Ticket {r[16]}"]
            if x and x != "Ticket None"
        )[:500]
    )
    nombre = sql_text_or_null(r[4], 200)

    producto_json = f"""jsonb_build_object(
      'nombre', {nombre},
      'sku', {sql_quote(sku)},
      'codigo_barras', {sql_quote(bc) if bc else 'NULL'},
      'categoria', 'GENERAL',
      'tipo', {sql_quote(tipo)},
      'descripcion', {desc},
      'costo', {sql_num_or_null(costo)},
      'precio', {sql_num_or_null(precio)},
      'stock_minimo', 5,
      'activo', true,
      'requiere_receta', false
    )"""

    label = f"-- {r[16]} L{r[0]} {str(r[4] or '')[:50]}"
    if bc:
        return f"""
{label}
do $$
declare v_pid bigint; v_lid bigint;
begin
  select id into v_pid from public.productos where codigo_barras = {sql_quote(bc)} limit 1;
  if v_pid is null then
    select producto_id into v_pid from _fc_carga_map where codigo_barras = {sql_quote(bc)};
  end if;
  if v_pid is null then
    select f.producto_id, f.lote_id into v_pid, v_lid
    from create_producto_with_lote(
      {producto_json},
      {qty},
      {sql_quote(lote)},
      {cad_sql},
      {sql_num_or_null(costo)},
      null,
      {proveedor_arg}
    ) f;
    insert into _fc_carga_map (codigo_barras, producto_id)
    values ({sql_quote(bc)}, v_pid)
    on conflict (codigo_barras) do update set producto_id = excluded.producto_id;
  else
    perform lote_id from receive_merchandise_lote(
      v_pid, {qty}, {sql_quote(lote)}, {cad_sql}, {sql_num_or_null(costo)}, {proveedor}, null
    );
  end if;
end $$;
"""
    return f"""
{label} (sin barcode)
select producto_id, lote_id from create_producto_with_lote(
  {producto_json},
  {qty},
  {sql_quote(lote)},
  {cad_sql},
  {sql_num_or_null(costo)},
  null,
  {proveedor_arg}
);
"""


def write_sql_files(rows: list[tuple]) -> list[Path]:
    SQL_DIR.mkdir(parents=True, exist_ok=True)
    stamp = date.today().strftime("%Y%m%d")
    base = f"carga_tickets_{stamp}"

    header = f"""-- FarmaCapital — Carga inventario tickets 2026-08-08
-- Filas: {len(rows)} | Margen venta: {int(MARGEN_VENTA * 100)}% sobre costo
-- 1) Backup  2) Opcional reset_pre_lanzamiento.sql  3) Ejecutar partes en orden
-- 4) patch_resync_productos_stock_from_lotes.sql

begin;

create temp table if not exists _fc_carga_map (
  codigo_barras text primary key,
  producto_id bigint
) on commit drop;

insert into _fc_carga_map (codigo_barras, producto_id)
select codigo_barras, id from public.productos
where codigo_barras is not null and btrim(codigo_barras) <> ''
on conflict (codigo_barras) do nothing;

"""
    footer = """
update public.productos p
set stock = coalesce((
  select sum(l.cantidad_actual)
  from public.lotes l
  where l.producto_id = p.id and coalesce(l.activo, true) = true
), 0);

commit;
"""

    parts: list[str] = []
    current = header
    for r in rows:
        block = sql_block_for_row(r)
        if len((current + block).encode("utf-8")) > MAX_SQL_BYTES:
            parts.append(current)
            current = (
                "begin;\n\ncreate temp table if not exists _fc_carga_map (\n"
                "  codigo_barras text primary key,\n  producto_id bigint\n) on commit drop;\n\n"
                "insert into _fc_carga_map (codigo_barras, producto_id)\n"
                "select codigo_barras, id from public.productos\n"
                "where codigo_barras is not null and btrim(codigo_barras) <> ''\n"
                "on conflict (codigo_barras) do nothing;\n\n"
            )
        current += block
    current += footer
    parts.append(current)

    paths: list[Path] = []
    for i, content in enumerate(parts, start=1):
        p = SQL_DIR / f"{base}_part{str(i).zfill(4)}.sql"
        p.write_text(content, encoding="utf-8")
        paths.append(p)

    manifest = SQL_DIR / f"{base}_manifest.txt"
    manifest.write_text(
        "\n".join(
            [
                f"stamp: {stamp}",
                f"filas: {len(rows)}",
                f"excel: {OUTPUT_XLSX}",
                "",
                *[x.name for x in paths],
                "",
                "Ejecutar cada part000N.sql en orden en Supabase SQL Editor.",
                "Luego: sql/patch_resync_productos_stock_from_lotes.sql",
            ]
        ),
        encoding="utf-8",
    )
    return paths


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"No existe fuente: {SOURCE_XLSX}")

    print("OCR desde PDFs (Vision macOS)...")
    ocr = load_ocr_from_pdfs(force=False)
    all_rows: list[tuple] = []

    existing = load_existing_maestro(SOURCE_XLSX)
    equilibrio = [r for r in existing if r[16] == "440393"]
    all_rows.extend(equilibrio)

    bodega = parse_bodega(ocr["Bodega F-42.pdf"])
    all_rows.extend(bodega)

    all_rows.extend(parse_surtidor(ocr["El surtidor de su farmacia.pdf"]))
    all_rows.extend(parse_ifc(ocr["IFC 1.pdf"], "IFC1-080826", "118217"))
    all_rows.extend(parse_ifc(ocr["IFC 2.pdf"], "IFC2-080826", "118216"))
    all_rows.extend(parse_farma_mx(ocr["Farma Mx.pdf"]))
    all_rows.extend(parse_farmalive(ocr["FarmaLive.pdf"]))

    stats = []
    by_ticket: dict[str, list[tuple]] = defaultdict(list)
    for r in all_rows:
        by_ticket[str(r[16])].append(r)
    for ticket, rs in sorted(by_ticket.items()):
        stats.append(
            {
                "ticket": ticket,
                "prov": rs[0][13],
                "lines": len(rs),
                "pieces": sum(int(x[8] or 0) for x in rs),
                "cost": round(sum(float(x[10] or 0) for x in rs), 2),
                "notes": rs[0][19],
            }
        )

    write_workbook(all_rows, stats)
    sql_paths = write_sql_files(all_rows)

    print(f"Generado: {OUTPUT_XLSX}")
    print(f"Total filas: {len(all_rows)}")
    for s in stats:
        print(
            f"  {s['ticket']} | {s['prov']} | líneas={s['lines']} | piezas={s['pieces']} | ${s['cost']}"
        )
    print(f"\nSQL ({len(sql_paths)} fragmentos):")
    for p in sql_paths:
        print(f"  {p}")


if __name__ == "__main__":
    main()
