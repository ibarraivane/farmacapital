"""Excel Comparativo_Multiproveedor_YYYYMMDD.xlsx."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lib.pricing.match import Match, ProductoCatalogo, alcanzable
from lib.pricing.normalize import precio_por_unidad_base

FONT = "Arial"
HEADER = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
NORMAL = Font(name=FONT, size=10)
VERDE = PatternFill("solid", fgColor="C6EFCE")   # A / confirmado
AMARILLO = PatternFill("solid", fgColor="FFF2CC")  # B
ROJO = PatternFill("solid", fgColor="FFC7CE")    # C
GRIS = PatternFill("solid", fgColor="D9D9D9")    # sin match
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

SEMAFORO = {
    "A": VERDE,
    "confirmado": VERDE,
    "B": AMARILLO,
    "C": ROJO,
    "ninguno": GRIS,
}

COLS_PROV = [
    "SKU", "Producto FC", "Marca", "Presentación", "Tu costo",
    "Producto proveedor", "Precio empaque", "Pzas empaque",
    "Precio unitario", "MXN / unidad base", "Nivel", "Score", "URL",
]
COLS_MEJOR = [
    "SKU", "Producto FC", "Marca", "Tu costo",
    "Proveedor ganador", "Producto proveedor", "Precio unitario",
    "MXN / unidad base", "Ahorro $", "Ahorro %", "Alcanzable", "Nivel",
]
COLS_CAND = [
    "SKU", "Producto FC", "Marca", "Fuente",
    "Candidato 1", "Score 1", "Candidato 2", "Score 2", "Candidato 3", "Score 3",
]


def _header(ws, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(1, i, h)
        c.font = HEADER_FONT
        c.fill = HEADER
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        c.border = THIN


def _cell(ws, r, col, value, fill=None):
    c = ws.cell(r, col, value)
    c.font = NORMAL
    c.border = THIN
    if fill is not None:
        c.fill = fill
    return c


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _base(cat: ProductoCatalogo, m: Match) -> float | None:
    if m.proveedor is None:
        return None
    if m.proveedor.precio_base is not None:
        return m.proveedor.precio_base
    return precio_por_unidad_base(m.proveedor.precio_unitario, m.proveedor.tamano or cat.tamano)


def escribir_comparativo(
    catalogo: list[ProductoCatalogo],
    por_fuente: dict[str, list[Match]],
    path: Path,
) -> Path:
    wb = Workbook()
    # una hoja por proveedor
    first = True
    for fuente, matches in por_fuente.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = fuente[:31]
        _header(ws, COLS_PROV)
        by_sku = {m.sku: m for m in matches}
        r = 2
        for cat in catalogo:
            m = by_sku.get(cat.sku) or Match(cat.sku, fuente, "ninguno", 0, None)
            fill = SEMAFORO.get(m.nivel, GRIS)
            p = m.proveedor
            vals = [
                cat.sku,
                cat.nombre,
                cat.marca,
                cat.presentacion,
                cat.costo,
                p.producto_raw if p else "",
                p.precio_empaque if p else "",
                p.cantidad_empaque if p else "",
                p.precio_unitario if p else "",
                _base(cat, m) if p else "",
                m.nivel if p or m.nivel == "C" else "",
                round(m.score, 1) if m.score else "",
                p.url if p else "",
            ]
            for i, v in enumerate(vals, start=1):
                _cell(ws, r, i, v, fill if i >= 6 else None)
            r += 1
        _widths(ws, [14, 36, 14, 14, 12, 40, 14, 12, 14, 16, 12, 8, 36])

    # Mejor_Precio — solo mayorista / distribuidor_farma, niveles A/B/confirmado
    ws = wb.create_sheet("Mejor_Precio")
    _header(ws, COLS_MEJOR)
    r = 2
    ahorro_total = 0.0
    cubiertos = 0
    for cat in catalogo:
        candidatos = []
        for fuente, matches in por_fuente.items():
            m = next((x for x in matches if x.sku == cat.sku), None)
            if not m or not m.proveedor:
                continue
            if m.proveedor.tipo_fuente == "benchmark_retail":
                continue
            if m.nivel not in {"A", "B", "confirmado"}:
                continue
            if m.proveedor.precio_unitario is None:
                continue
            candidatos.append(m)
        if not candidatos:
            vals = [cat.sku, cat.nombre, cat.marca, cat.costo, "", "", "", "", "", "", "", ""]
            for i, v in enumerate(vals, start=1):
                _cell(ws, r, i, v, GRIS)
            r += 1
            continue
        best = min(candidatos, key=lambda m: m.proveedor.precio_unitario)
        p = best.proveedor
        costo = cat.costo
        ahorro = (costo - p.precio_unitario) if (costo is not None and p.precio_unitario is not None) else None
        pct = (ahorro / costo * 100) if (ahorro is not None and costo) else None
        if ahorro is not None and ahorro > 0:
            ahorro_total += ahorro
            cubiertos += 1
        alcan = alcanzable(p)
        fill = VERDE if (ahorro or 0) > 0 else (AMARILLO if best.nivel == "B" else GRIS)
        vals = [
            cat.sku, cat.nombre, cat.marca, costo,
            p.fuente, p.producto_raw, p.precio_unitario, _base(cat, best),
            round(ahorro, 2) if ahorro is not None else "",
            round(pct, 1) if pct is not None else "",
            "sí" if alcan else "no",
            best.nivel,
        ]
        for i, v in enumerate(vals, start=1):
            _cell(ws, r, i, v, fill)
        r += 1
    _widths(ws, [14, 36, 14, 12, 16, 40, 14, 16, 12, 10, 12, 12])

    # Candidatos_Revisión — solo nivel C
    ws = wb.create_sheet("Candidatos_Revisión")
    _header(ws, COLS_CAND)
    r = 2
    for fuente, matches in por_fuente.items():
        for m in matches:
            if m.nivel != "C" or not m.candidatos:
                continue
            cat = next((c for c in catalogo if c.sku == m.sku), None)
            if cat is None:
                continue
            cands = m.candidatos[:3]
            vals = [cat.sku, cat.nombre, cat.marca, fuente]
            for p, sc in cands:
                vals.extend([p.producto_raw, round(sc, 1)])
            while len(vals) < 10:
                vals.append("")
            for i, v in enumerate(vals, start=1):
                _cell(ws, r, i, v, ROJO if i >= 5 else None)
            r += 1
    _widths(ws, [14, 36, 14, 14, 40, 10, 40, 10, 40, 10])

    # Resumen: cobertura importa tanto como el precio
    ws = wb.create_sheet("Resumen")
    _header(ws, [
        "Proveedor", "Tipo", "SKUs cubiertos (A/B/confirmado)",
        "Cobertura %", "Matches A", "Matches B", "Candidatos C",
        "Ahorro potencial $ (filas alcanzables)",
    ])
    n_cat = len(catalogo) or 1
    r = 2
    for fuente, matches in por_fuente.items():
        tipo = ""
        a = b = c = cub = 0
        ahorro = 0.0
        by_sku = {m.sku: m for m in matches}
        for cat in catalogo:
            m = by_sku.get(cat.sku)
            if not m:
                continue
            if m.proveedor:
                tipo = m.proveedor.tipo_fuente
            if m.nivel == "A" or m.nivel == "confirmado":
                a += 1
                cub += 1
            elif m.nivel == "B":
                b += 1
                cub += 1
            elif m.nivel == "C":
                c += 1
            if m.nivel in {"A", "B", "confirmado"} and m.proveedor and m.proveedor.tipo_fuente != "benchmark_retail":
                if cat.costo and m.proveedor.precio_unitario is not None and alcanzable(m.proveedor):
                    delta = cat.costo - m.proveedor.precio_unitario
                    if delta > 0:
                        ahorro += delta
        vals = [fuente, tipo, cub, round(100 * cub / n_cat, 1), a, b, c, round(ahorro, 2)]
        for i, v in enumerate(vals, start=1):
            _cell(ws, r, i, v)
        r += 1
    ws.cell(r + 1, 1, "Nota: un proveedor 3% más caro que cubre el 70% del catálogo vale más que uno 8% más barato al 5%.")
    ws.cell(r + 2, 1, f"Catálogo: {len(catalogo)} SKUs. Fecha: {date.today().isoformat()}.")
    _widths(ws, [18, 20, 28, 12, 12, 12, 14, 32])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
