"""Lectura / escritura del CSV normalizado y del catálogo FarmaCapital."""

from __future__ import annotations

import csv
from pathlib import Path

from lib.pricing.schema import COLUMNAS_CSV, fila_vacia

ROOT = Path(__file__).resolve().parents[2]
DIR_PROVEEDORES = ROOT / "pricing" / "precios_proveedores"
CATALOGO_CSV = ROOT / "sql" / "preview_catalogo_campos_y_precios.csv"
CATALOGO_XLSX = ROOT / "Inventario_FarmaCapital.xlsx"


def escribir_csv(path: Path, filas: list[dict]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNAS_CSV, extrasaction="ignore")
        w.writeheader()
        for f in filas:
            row = fila_vacia(**{k: f.get(k) for k in COLUMNAS_CSV})
            w.writerow(row)
    return path


def leer_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def _num(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def cargar_catalogo(path: Path | None = None) -> list[dict]:
    """Catálogo FarmaCapital.

    Preferimos el preview CSV (el mismo que usa pricing_pipeline.py): el xlsx
    local a veces trae filas de título y queda desfasado. Si se pasa un path,
    se respeta. El xlsx se usa como fallback y se busca la fila con 'SKU'.
    """
    if path is not None:
        return _catalogo_xlsx(path) if str(path).endswith(".xlsx") else _catalogo_csv(path)
    if CATALOGO_CSV.exists():
        return _catalogo_csv(CATALOGO_CSV)
    if CATALOGO_XLSX.exists():
        return _catalogo_xlsx(CATALOGO_XLSX)
    return []


def _catalogo_csv(path: Path) -> list[dict]:
    rows = []
    with path.open(encoding="utf-8", newline="") as fh:
        for r in csv.DictReader(fh):
            if str(r.get("activo", "True")).lower() == "false":
                continue
            rows.append({
                "sku": (r.get("sku") or "").strip(),
                "nombre": (r.get("nombre") or "").strip(),
                "nombre_original": (r.get("denominacion_distintiva") or r.get("nombre") or "").strip(),
                "marca": (r.get("marca") or "").strip(),
                "presentacion": (r.get("presentacion") or "").strip(),
                "principio_activo": (r.get("principio_activo") or "").strip(),
                "tipo": (r.get("tipo") or r.get("categoria") or "").strip(),
                "costo": _num(r.get("costo")),
                "precio_venta": _num(r.get("precio") or r.get("calculated_price")),
            })
    return [r for r in rows if r["sku"]]


def _catalogo_xlsx(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("Falta openpyxl para leer el inventario xlsx") from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Inventario"] if "Inventario" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = None
    for raw in rows_iter:
        cells = [str(h or "").strip() for h in raw]
        if any(c.lower() == "sku" for c in cells):
            header = cells
            break
    if not header:
        return []
    alias = {h.lower(): i for i, h in enumerate(header)}

    def col(*nombres):
        for n in nombres:
            if n.lower() in alias:
                return alias[n.lower()]
        return None

    i_sku = col("SKU", "sku")
    i_nom = col("Nombre", "nombre")
    i_orig = col("Nombre original", "nombre_original")
    i_marca = col("Marca", "marca")
    i_pres = col("Presentación", "Presentacion", "presentacion")
    i_pa = col("Principio activo", "principio_activo")
    i_tipo = col("Tipo", "tipo")
    i_costo = col("Costo", "costo")
    i_precio = col("Precio Venta", "precio", "precio_venta")
    out = []
    for row in rows_iter:
        sku = str(row[i_sku] or "").strip() if i_sku is not None else ""
        if not sku:
            continue
        out.append({
            "sku": sku,
            "nombre": str(row[i_nom] or "").strip() if i_nom is not None else "",
            "nombre_original": str(row[i_orig] or row[i_nom] or "").strip() if (i_orig or i_nom) is not None else "",
            "marca": str(row[i_marca] or "").strip() if i_marca is not None else "",
            "presentacion": str(row[i_pres] or "").strip() if i_pres is not None else "",
            "principio_activo": str(row[i_pa] or "").strip() if i_pa is not None else "",
            "tipo": str(row[i_tipo] or "").strip() if i_tipo is not None else "",
            "costo": _num(row[i_costo]) if i_costo is not None else None,
            "precio_venta": _num(row[i_precio]) if i_precio is not None else None,
        })
    return out
